"""Tests für das Konfigurationssystem und Phase-1-Datenmodelle."""

import tempfile
from pathlib import Path

import pytest

from config.schema import (
    DoubleBlock,
    GradeConfig,
    GradeDefinition,
    LessonSlot,
    PauseSlot,
    SchoolConfig,
    SchoolType,
    TimeGridConfig,
)
from config.defaults import (
    default_grades,
    default_rooms,
    default_school_config,
    default_time_grid,
    STUNDENTAFEL_GYMNASIUM_SEK1,
    SUBJECT_METADATA,
)
from config.manager import ConfigManager


# ─── DEFAULT-KONFIGURATION ────────────────────────────────────────────────────

class TestDefaultConfig:
    def test_default_time_grid_valid(self):
        """Default-Zeitraster lässt sich ohne Fehler erstellen."""
        tg = default_time_grid()
        assert tg.days_per_week == 5
        assert len(tg.lesson_slots) == 10
        assert len(tg.pauses) == 3
        assert len(tg.double_blocks) == 3

    def test_default_grades_valid(self):
        """Default-Jahrgangskonfiguration ist korrekt."""
        gc = default_grades()
        assert len(gc.grades) == 6
        assert gc.total_classes == 36
        assert gc.grade_numbers == [5, 6, 7, 8, 9, 10]

    def test_default_rooms_valid(self):
        """Default-Fachräume sind korrekt."""
        rc = default_rooms()
        assert len(rc.special_rooms) == 7
        physik = next(r for r in rc.special_rooms if r.room_type == "physik")
        assert physik.count == 3
        bio = next(r for r in rc.special_rooms if r.room_type == "biologie")
        assert bio.count == 3  # 36 Klassen brauchen mind. 3 Bio-Räume

    def test_default_school_config_valid(self):
        """Vollständige Default-Config ist valide."""
        config = default_school_config()
        assert config.school_name == "Muster-Gymnasium"
        assert config.school_type == SchoolType.GYMNASIUM
        assert config.bundesland == "NRW"
        assert config.grades.total_classes == 36
        assert config.teachers.total_count == 60

    def test_room_get_capacity_special(self):
        """get_capacity gibt korrekte Kapazität für Fachräume zurück."""
        rc = default_rooms()
        assert rc.get_capacity("physik") == 3
        assert rc.get_capacity("chemie") == 2

    def test_room_get_capacity_unknown(self):
        """get_capacity gibt 999 für unbekannte Raumtypen zurück."""
        rc = default_rooms()
        assert rc.get_capacity("unbekannt") == 999

    def test_stundentafel_covers_all_grades(self):
        """Stundentafel enthält alle Jahrgänge 5-10."""
        gc = default_grades()
        for g in gc.grade_numbers:
            assert g in STUNDENTAFEL_GYMNASIUM_SEK1, f"Jahrgang {g} fehlt in Stundentafel"

    def test_subject_metadata_not_empty(self):
        """SUBJECT_METADATA enthält alle relevanten Fächer."""
        required = ["Deutsch", "Mathematik", "Englisch", "Physik",
                    "Chemie", "Biologie", "Sport"]
        for subj in required:
            assert subj in SUBJECT_METADATA, f"Fach '{subj}' fehlt in SUBJECT_METADATA"


# ─── PYDANTIC-VALIDIERUNG ─────────────────────────────────────────────────────

class TestPydanticValidation:
    def test_valid_double_block(self):
        """Gültige Doppelstunden-Blöcke werden akzeptiert."""
        tg = TimeGridConfig(
            lesson_slots=[
                LessonSlot(slot_number=1, start_time="08:00", end_time="08:45"),
                LessonSlot(slot_number=2, start_time="08:45", end_time="09:30"),
            ],
            pauses=[],
            double_blocks=[DoubleBlock(slot_first=1, slot_second=2)],
        )
        assert len(tg.double_blocks) == 1

    def test_double_block_over_pause_raises(self):
        """Doppelstunden-Block über Pause hinweg → Validierungsfehler."""
        with pytest.raises(Exception):
            TimeGridConfig(
                lesson_slots=[
                    LessonSlot(slot_number=1, start_time="08:00", end_time="08:45"),
                    LessonSlot(slot_number=2, start_time="09:00", end_time="09:45"),
                ],
                pauses=[PauseSlot(after_slot=1, duration_minutes=15)],
                double_blocks=[DoubleBlock(slot_first=1, slot_second=2)],
            )

    def test_double_block_nonsequential_raises(self):
        """Nicht aufeinanderfolgende Blöcke → Validierungsfehler."""
        with pytest.raises(Exception):
            TimeGridConfig(
                lesson_slots=[
                    LessonSlot(slot_number=1, start_time="08:00", end_time="08:45"),
                    LessonSlot(slot_number=2, start_time="08:45", end_time="09:30"),
                    LessonSlot(slot_number=3, start_time="09:30", end_time="10:15"),
                ],
                pauses=[],
                double_blocks=[DoubleBlock(slot_first=1, slot_second=3)],
            )

    def test_double_block_missing_slot_raises(self):
        """Block referenziert nicht-existenten Slot → Validierungsfehler."""
        with pytest.raises(Exception):
            TimeGridConfig(
                lesson_slots=[
                    LessonSlot(slot_number=1, start_time="08:00", end_time="08:45"),
                ],
                pauses=[],
                double_blocks=[DoubleBlock(slot_first=1, slot_second=2)],
            )

    def test_teacher_config_defaults_valid(self):
        """TeacherConfig mit Defaults ist valide."""
        from config.schema import TeacherConfig
        tc = TeacherConfig()
        assert tc.total_count == 60
        assert tc.vollzeit_deputat == 26

    def test_grade_total_classes(self):
        """GradeConfig.total_classes berechnet korrekt."""
        gc = GradeConfig(grades=[
            GradeDefinition(grade=5, num_classes=6, weekly_hours_target=30),
            GradeDefinition(grade=6, num_classes=5, weekly_hours_target=31),
        ])
        assert gc.total_classes == 11

    def test_invalid_days_per_week(self):
        """Ungültige Tagesanzahl (< 5 oder > 6) → Validierungsfehler."""
        with pytest.raises(Exception):
            TimeGridConfig(
                days_per_week=4,
                lesson_slots=[LessonSlot(slot_number=1,
                                         start_time="08:00", end_time="08:45")],
                pauses=[],
                double_blocks=[],
            )


# ─── YAML SPEICHERN / LADEN ───────────────────────────────────────────────────

class TestConfigManager:
    def test_save_and_load_roundtrip(self, tmp_path: Path):
        """Config speichern, laden und validieren — vollständiger Roundtrip."""
        config = default_school_config()
        mgr = ConfigManager()
        mgr.CONFIG_DIR = tmp_path
        mgr.DEFAULT_CONFIG = tmp_path / "school_config.yaml"

        mgr.save(config)
        assert mgr.DEFAULT_CONFIG.exists()

        loaded = mgr.load(mgr.DEFAULT_CONFIG)
        assert loaded.school_name == config.school_name
        assert loaded.grades.total_classes == config.grades.total_classes
        assert loaded.teachers.total_count == config.teachers.total_count

    def test_first_run_check_no_file(self, tmp_path: Path):
        """first_run_check gibt True zurück wenn keine Config existiert."""
        mgr = ConfigManager()
        mgr.DEFAULT_CONFIG = tmp_path / "nonexistent.yaml"
        assert mgr.first_run_check() is True

    def test_first_run_check_with_file(self, tmp_path: Path):
        """first_run_check gibt False zurück wenn Config existiert."""
        config = default_school_config()
        mgr = ConfigManager()
        mgr.CONFIG_DIR = tmp_path
        mgr.DEFAULT_CONFIG = tmp_path / "school_config.yaml"
        mgr.save(config)
        assert mgr.first_run_check() is False

    def test_load_nonexistent_raises(self, tmp_path: Path):
        """Laden einer nicht-existenten Datei → FileNotFoundError."""
        mgr = ConfigManager()
        with pytest.raises(FileNotFoundError):
            mgr.load(tmp_path / "not_there.yaml")

    def test_scenario_save_and_load(self, tmp_path: Path):
        """Szenario speichern und laden — Roundtrip."""
        config = default_school_config()
        config = config.model_copy(update={"school_name": "Test-Schule"})

        mgr = ConfigManager()
        mgr.CONFIG_DIR = tmp_path
        mgr.DEFAULT_CONFIG = tmp_path / "school_config.yaml"
        mgr.SCENARIOS_DIR = tmp_path / "scenarios"

        mgr.save_scenario(config, "test_szenario", "Nur zum Testen")
        loaded = mgr.load_scenario("test_szenario")
        assert loaded.school_name == "Test-Schule"

    def test_list_scenarios_empty(self, tmp_path: Path):
        """Leeres Szenario-Verzeichnis → leere Liste."""
        mgr = ConfigManager()
        mgr.SCENARIOS_DIR = tmp_path / "scenarios"
        assert mgr.list_scenarios() == []


# ─── MODELLE (Phase 1: Pydantic v2) ──────────────────────────────────────────

class TestModels:
    def test_subject_pydantic(self):
        """Subject ist ein Pydantic-Modell mit korrekten Feldern."""
        from models.subject import Subject
        s = Subject(name="Mathematik", short_name="Ma", category="hauptfach",
                    is_hauptfach=True)
        assert s.name == "Mathematik"
        assert s.short_name == "Ma"
        assert not s.needs_special_room

    def test_subject_with_room(self):
        """Subject mit Fachraum: needs_special_room=True."""
        from models.subject import Subject
        s = Subject(name="Physik", short_name="Ph", category="nw",
                    requires_special_room="physik", double_lesson_required=True)
        assert s.needs_special_room
        assert s.requires_special_room == "physik"

    def test_teacher_pydantic(self):
        """Teacher ist ein Pydantic-Modell; id wird normalisiert."""
        from models.teacher import Teacher
        t = Teacher(id="mül", name="Müller, Hans", subjects=["Mathematik", "Physik"],
                    deputat_max=26, deputat_min=21)
        assert t.id == "MÜL"  # normalisiert zu Großbuchstaben
        assert t.deputat == 26  # Property gibt deputat_max zurück
        assert t.deputat_max == 26
        assert t.deputat_min == 21
        assert not t.is_teilzeit

    def test_teacher_unavailable_slots(self):
        """Teacher.unavailable_slots akzeptiert tuple[int,int]."""
        from models.teacher import Teacher
        t = Teacher(id="TST", name="Test, A", subjects=["Deutsch"], deputat_max=20, deputat_min=16,
                    unavailable_slots=[(0, 1), (0, 2), (4, 7)])
        assert len(t.unavailable_slots) == 3

    def test_school_class_pydantic(self):
        """SchoolClass hat id, curriculum und max_slot."""
        from models.school_class import SchoolClass
        sc = SchoolClass(
            id="7b", grade=7, label="b",
            curriculum={"Mathematik": 4, "Deutsch": 4, "Englisch": 3},
            max_slot=7,
        )
        assert sc.id == "7b"
        assert sc.total_weekly_hours == 11
        assert sc.max_slot == 7

    def test_room_pydantic(self):
        """Room hat id, room_type, name."""
        from models.room import Room
        r = Room(id="PH1", room_type="physik", name="Physik-Raum 1")
        assert r.room_type == "physik"

    def test_coupling_group(self):
        """CouplingGroup hat group_name, subject, hours_per_week."""
        from models.coupling import CouplingGroup
        cg = CouplingGroup(group_name="evangelisch", subject="Religion", hours_per_week=2)
        assert cg.group_name == "evangelisch"

    def test_coupling_pydantic(self):
        """Coupling hat id, coupling_type, involved_class_ids, groups."""
        from models.coupling import Coupling, CouplingGroup
        c = Coupling(
            id="reli_5",
            coupling_type="reli_ethik",
            involved_class_ids=["5a", "5b", "5c"],
            groups=[
                CouplingGroup(group_name="evangelisch", subject="Religion", hours_per_week=2),
                CouplingGroup(group_name="ethik", subject="Ethik", hours_per_week=2),
            ],
            hours_per_week=2,
        )
        assert c.id == "reli_5"
        assert len(c.groups) == 2
        assert c.cross_class is True  # Default


# ─── FAKE-DATEN ───────────────────────────────────────────────────────────────

class TestFakeData:
    def test_generate_returns_school_data(self):
        """FakeDataGenerator.generate() gibt SchoolData zurück."""
        from data.fake_data import FakeDataGenerator
        from models.school_data import SchoolData
        config = default_school_config()
        gen = FakeDataGenerator(config, seed=42)
        data = gen.generate()
        assert isinstance(data, SchoolData)

    def test_generate_all_compat(self):
        """generate_all() (compat) gibt dict zurück."""
        from data.fake_data import FakeDataGenerator
        config = default_school_config()
        gen = FakeDataGenerator(config, seed=42)
        data = gen.generate_all()
        assert isinstance(data, dict)
        assert len(data["subjects"]) > 0
        assert len(data["classes"]) == 36
        assert len(data["teachers"]) == config.teachers.total_count

    def test_generate_classes_count(self):
        """Anzahl generierter Klassen stimmt mit Config überein."""
        from data.fake_data import FakeDataGenerator
        config = default_school_config()
        gen = FakeDataGenerator(config, seed=0)
        data = gen.generate()
        assert len(data.classes) == config.grades.total_classes

    def test_generate_teachers_count(self):
        """Anzahl generierter Lehrer stimmt mit Config überein."""
        from data.fake_data import FakeDataGenerator
        config = default_school_config()
        gen = FakeDataGenerator(config, seed=0)
        data = gen.generate()
        assert len(data.teachers) == config.teachers.total_count

    def test_teacher_ids_unique(self):
        """Lehrer-IDs (Kürzel) sind eindeutig."""
        from data.fake_data import FakeDataGenerator
        config = default_school_config()
        gen = FakeDataGenerator(config, seed=1)
        data = gen.generate()
        ids = [t.id for t in data.teachers]
        assert len(ids) == len(set(ids)), "Doppelte Lehrer-Kürzel gefunden!"

    def test_generate_rooms(self):
        """Fachräume werden korrekt generiert (Physik:3, Chemie:2, Bio:3, ...)."""
        from data.fake_data import FakeDataGenerator
        config = default_school_config()
        gen = FakeDataGenerator(config, seed=0)
        data = gen.generate()
        # Physik:3, Chemie:2, Bio:3, Inf:2, Kunst:3, Musik:3, Sport:4 = 20
        assert len(data.rooms) == 20

    def test_classes_have_curriculum(self):
        """Alle Klassen haben ein nicht-leeres Curriculum."""
        from data.fake_data import FakeDataGenerator
        config = default_school_config()
        gen = FakeDataGenerator(config, seed=42)
        data = gen.generate()
        for cls in data.classes:
            assert len(cls.curriculum) > 0, f"Klasse {cls.id} hat leeres Curriculum"
            assert cls.total_weekly_hours > 0

    def test_classes_have_max_slot(self):
        """Alle Klassen haben max_slot aus Config."""
        from data.fake_data import FakeDataGenerator
        config = default_school_config()
        gen = FakeDataGenerator(config, seed=42)
        data = gen.generate()
        expected_max = config.time_grid.sek1_max_slot
        for cls in data.classes:
            assert cls.max_slot == expected_max

    def test_chemie_engpass_bottleneck(self):
        """Chemie-Engpass: Nur 2 Chemie-Lehrkräfte vorhanden."""
        from data.fake_data import FakeDataGenerator
        config = default_school_config()
        gen = FakeDataGenerator(config, seed=42)
        data = gen.generate()
        chemie_teachers = [t for t in data.teachers if "Chemie" in t.subjects]
        assert len(chemie_teachers) == 2, (
            f"Erwartet 2 Chemie-Lehrkräfte, gefunden: {len(chemie_teachers)}"
        )

    def test_freitag_cluster_bottleneck(self):
        """Freitag-Cluster: Mindestens 4 Lehrkräfte wünschen Freitag frei."""
        from data.fake_data import FakeDataGenerator
        config = default_school_config()
        gen = FakeDataGenerator(config, seed=42)
        data = gen.generate()
        fr_cluster = [t for t in data.teachers if 4 in t.preferred_free_days]
        assert len(fr_cluster) >= 4

    def test_restricted_teacher_bottleneck(self):
        """Stark eingeschränkter Lehrer: Hat Sperrzeiten für Mo, Fr und Di-Nachmittag."""
        from data.fake_data import FakeDataGenerator
        config = default_school_config()
        gen = FakeDataGenerator(config, seed=42)
        data = gen.generate()
        # Suche Lehrer mit vielen gesperrten Slots (≥ 18 = Mo7 + Fr7 + Di-Nachmittag4)
        restricted = [t for t in data.teachers if len(t.unavailable_slots) >= 18]
        assert len(restricted) >= 1, "Kein eingeschränkter Lehrer gefunden!"

    def test_generate_subjects_from_metadata(self):
        """Alle Fächer aus SUBJECT_METADATA werden erzeugt."""
        from data.fake_data import FakeDataGenerator
        config = default_school_config()
        gen = FakeDataGenerator(config, seed=0)
        data = gen.generate()
        names = {s.name for s in data.subjects}
        assert "Deutsch" in names
        assert "Mathematik" in names
        assert "Physik" in names

    def test_couplings_generated(self):
        """Kopplungen (Reli/Ethik und WPF) werden erzeugt."""
        from data.fake_data import FakeDataGenerator
        config = default_school_config()
        gen = FakeDataGenerator(config, seed=42)
        data = gen.generate()
        assert len(data.couplings) > 0
        reli = [c for c in data.couplings if c.coupling_type == "reli_ethik"]
        wpf = [c for c in data.couplings if c.coupling_type == "wpf"]
        assert len(reli) == 6  # Jahrgänge 5-10
        assert len(wpf) == 2   # Jahrgänge 9-10


# ─── SCHOOL DATA + FEASIBILITY ────────────────────────────────────────────────

class TestSchoolData:
    def _make_data(self, seed: int = 42):
        from data.fake_data import FakeDataGenerator
        config = default_school_config()
        gen = FakeDataGenerator(config, seed=seed)
        return gen.generate()

    def test_summary_contains_key_info(self):
        """summary() enthält Schlüsselinfos."""
        data = self._make_data()
        s = data.summary()
        assert "Klassen" in s
        assert "Lehrkräfte" in s
        assert "Gesamtdeputat" in s

    def test_validate_feasibility_returns_report(self):
        """validate_feasibility() gibt FeasibilityReport zurück."""
        from models.school_data import FeasibilityReport
        data = self._make_data()
        report = data.validate_feasibility()
        assert isinstance(report, FeasibilityReport)
        assert isinstance(report.is_feasible, bool)
        assert isinstance(report.errors, list)
        assert isinstance(report.warnings, list)

    def test_chemie_engpass_triggers_warning(self):
        """Chemie-Engpass erzeugt Warnung wenn kein Mehrarbeit-Puffer gesetzt ist."""
        from data.fake_data import FakeDataGenerator
        config = default_school_config()
        # Ohne Puffer: Chemie-Kapazität 52h vs. 48h Bedarf = 108% → Warnung bei < 110%
        config = config.model_copy(update={
            "teachers": config.teachers.model_copy(update={"deputat_max_buffer": 0})
        })
        data = FakeDataGenerator(config, seed=42).generate()
        report = data.validate_feasibility()
        chemie_warnings = [w for w in report.warnings if "Chemie" in w or "chemie" in w.lower()]
        assert len(chemie_warnings) >= 1, (
            f"Keine Chemie-Warnung im Report. Warnungen: {report.warnings}"
        )

    def test_freitag_cluster_triggers_warning(self):
        """Freitag-Cluster erzeugt Warnung im Feasibility-Report."""
        data = self._make_data()
        report = data.validate_feasibility()
        fr_warnings = [w for w in report.warnings
                       if "Freitag" in w or "freitag" in w.lower()]
        assert len(fr_warnings) >= 1

    def test_feasibility_broken_data(self):
        """Kaputte Daten (kein Lehrer für Mathematik) → error im Report."""
        from models.school_data import SchoolData
        from models.teacher import Teacher
        from models.school_class import SchoolClass
        from models.subject import Subject

        config = default_school_config()
        data = SchoolData(
            subjects=[Subject(name="Mathematik", short_name="Ma", category="hauptfach")],
            rooms=[],
            classes=[SchoolClass(
                id="5a", grade=5, label="a",
                curriculum={"Mathematik": 4}, max_slot=7,
            )],
            teachers=[],  # Kein Lehrer!
            couplings=[],
            config=config,
        )
        report = data.validate_feasibility()
        assert not report.is_feasible
        assert any("Mathematik" in e for e in report.errors)

    def test_save_and_load_json(self, tmp_path: Path):
        """SchoolData → JSON → SchoolData Roundtrip."""
        data = self._make_data()
        json_path = tmp_path / "school_data.json"
        data.save_json(json_path)
        assert json_path.exists()

        loaded = type(data).load_json(json_path)
        assert len(loaded.classes) == len(data.classes)
        assert len(loaded.teachers) == len(data.teachers)
        assert loaded.config.school_name == data.config.school_name

    def test_load_json_nonexistent_raises(self, tmp_path: Path):
        """load_json mit nicht-existenter Datei → FileNotFoundError."""
        from models.school_data import SchoolData
        with pytest.raises(FileNotFoundError):
            SchoolData.load_json(tmp_path / "does_not_exist.json")


# ─── EXCEL TEMPLATE ───────────────────────────────────────────────────────────

class TestExcelTemplate:
    def test_template_creates_file(self, tmp_path: Path):
        """generate_template() erzeugt eine Excel-Datei."""
        from data.excel_import import generate_template
        config = default_school_config()
        out = tmp_path / "vorlage.xlsx"
        generate_template(config, out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_template_has_correct_sheets(self, tmp_path: Path):
        """Excel-Vorlage enthält alle 7 Pflicht-Blätter (inkl. Fächer)."""
        import openpyxl
        from data.excel_import import generate_template
        config = default_school_config()
        out = tmp_path / "vorlage.xlsx"
        generate_template(config, out)

        wb = openpyxl.load_workbook(str(out))
        sheet_names_lower = [s.lower() for s in wb.sheetnames]
        for expected in ["zeitraster", "jahrgänge", "fächer", "stundentafel",
                         "lehrkräfte", "fachräume", "kopplungen"]:
            assert expected in sheet_names_lower, \
                f"Blatt '{expected}' fehlt. Vorhanden: {wb.sheetnames}"

    def test_faecher_sheet_has_columns(self, tmp_path: Path):
        """Fächer-Blatt enthält die erwarteten Spaltenköpfe."""
        import openpyxl
        from data.excel_import import generate_template
        config = default_school_config()
        out = tmp_path / "vorlage.xlsx"
        generate_template(config, out)

        wb = openpyxl.load_workbook(str(out))
        ws = wb["Fächer"]
        headers = [str(c.value or "").strip().lower() for c in ws[1]]
        assert "fachname" in headers
        assert "kürzel" in headers
        assert "kategorie" in headers
        assert "hauptfach (ja/nein)" in headers

    def test_faecher_sheet_prefilled_from_metadata(self, tmp_path: Path):
        """Fächer-Blatt ist mit SUBJECT_METADATA vorausgefüllt."""
        import openpyxl
        from data.excel_import import generate_template
        config = default_school_config()
        out = tmp_path / "vorlage.xlsx"
        generate_template(config, out)

        wb = openpyxl.load_workbook(str(out))
        ws = wb["Fächer"]
        rows = list(ws.iter_rows(values_only=True))
        # Zeile 1 = Header, Zeile 2 = Hinweis-Merge, Zeilen 3+ = Fächer
        data_rows = [r for r in rows[2:] if any(v for v in r)]
        assert len(data_rows) >= len(SUBJECT_METADATA)
        first_name = str(data_rows[0][0] or "")
        assert first_name in SUBJECT_METADATA

    def test_stundentafel_sheet_has_note(self, tmp_path: Path):
        """Stundentafel-Blatt hat Hinweis-Zeile in Zeile 1."""
        import openpyxl
        from data.excel_import import generate_template
        config = default_school_config()
        out = tmp_path / "vorlage.xlsx"
        generate_template(config, out)

        wb = openpyxl.load_workbook(str(out))
        ws = wb["Stundentafel"]
        note = str(ws.cell(row=1, column=1).value or "").lower()
        assert "wochenstunden" in note or "fächer" in note

    def test_zeitraster_sheet_prefilled(self, tmp_path: Path):
        """Zeitraster-Blatt enthält vorausgefüllte Daten aus Config."""
        import openpyxl
        from data.excel_import import generate_template
        config = default_school_config()
        out = tmp_path / "vorlage.xlsx"
        generate_template(config, out)

        wb = openpyxl.load_workbook(str(out))
        ws = wb["Zeitraster"]
        rows = list(ws.iter_rows(values_only=True))
        # Header + mind. eine Datenzeile
        assert len(rows) >= 2
        # Erste Spalte der zweiten Zeile = Slot-Nummer 1
        assert rows[1][0] == 1


# ─── EXCEL IMPORT ─────────────────────────────────────────────────────────────

class TestExcelImport:
    """Tests für den ExcelImporter: Fächer, Stundentafel, Rückwärtskompatibilität."""

    def _make_template(self, tmp_path: Path) -> Path:
        from data.excel_import import generate_template
        config = default_school_config()
        out = tmp_path / "vorlage.xlsx"
        generate_template(config, out)
        return out

    def test_import_subjects_from_faecher_sheet(self, tmp_path: Path):
        """Fächer aus 'Fächer'-Blatt werden korrekt importiert."""
        import openpyxl
        from data.excel_import import ExcelImporter
        config = default_school_config()

        # Erstelle minimale Excel-Datei mit Fächer-Blatt
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fächer"
        ws.append(["Fachname", "Kürzel", "Kategorie", "Hauptfach (ja/nein)",
                   "Fachraum-Typ", "Doppelstunde Pflicht", "Doppelstunde Bevorzugt"])
        ws.append(["Türkisch", "TÜR", "sprachen", "nein", "", "nein", "nein"])
        ws.append(["Mathematik", "Ma", "hauptfach", "ja", "", "nein", "ja"])
        path = tmp_path / "faecher.xlsx"
        wb.save(str(path))

        importer = ExcelImporter(path, config)
        importer._open()
        subjects = importer.import_subjects()
        names = [s.name for s in subjects]
        assert "Türkisch" in names
        assert "Mathematik" in names
        assert len(subjects) == 2

    def test_import_subjects_fallback_when_sheet_absent(self, tmp_path: Path):
        """Kein 'Fächer'-Blatt → Fallback auf SUBJECT_METADATA."""
        import openpyxl
        from data.excel_import import ExcelImporter
        config = default_school_config()

        wb = openpyxl.Workbook()
        wb.active.title = "Lehrkräfte"
        path = tmp_path / "no_faecher.xlsx"
        wb.save(str(path))

        importer = ExcelImporter(path, config)
        importer._open()
        subjects = importer.import_subjects()
        names = [s.name for s in subjects]
        assert "Deutsch" in names
        assert "Mathematik" in names
        assert len(subjects) == len(SUBJECT_METADATA)

    def test_import_stundentafel_custom_values(self, tmp_path: Path):
        """Benutzerdefinierte Stundentafel wird korrekt importiert."""
        import openpyxl
        from data.excel_import import ExcelImporter
        config = default_school_config()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stundentafel"
        # Hinweis-Zeile (wird übersprungen)
        ws.append(["Wochenstunden pro Klasse und Fach"])
        # Kopfzeile
        ws.append(["Fach", "Jg. 5", "Jg. 6"])
        # Deutsch: Bayern hätte z.B. 5h in Jg.5 statt 4h
        ws.append(["Deutsch", 5, 4])
        ws.append(["Mathematik", 4, 4])
        path = tmp_path / "stundentafel.xlsx"
        wb.save(str(path))

        importer = ExcelImporter(path, config)
        importer._open()
        known = ["Deutsch", "Mathematik"]
        st = importer.import_stundentafel(known)
        assert st[5]["Deutsch"] == 5  # Bayern-Wert
        assert st[6]["Deutsch"] == 4
        assert st[5]["Mathematik"] == 4

    def test_import_stundentafel_fallback_when_absent(self, tmp_path: Path):
        """Kein 'Stundentafel'-Blatt → Fallback auf NRW-Stundentafel."""
        import openpyxl
        from data.excel_import import ExcelImporter
        config = default_school_config()

        wb = openpyxl.Workbook()
        wb.active.title = "Lehrkräfte"
        path = tmp_path / "no_st.xlsx"
        wb.save(str(path))

        importer = ExcelImporter(path, config)
        importer._open()
        st = importer.import_stundentafel(["Deutsch"])
        assert st == STUNDENTAFEL_GYMNASIUM_SEK1

    def test_import_stundentafel_missing_cell_is_zero(self, tmp_path: Path):
        """Fehlende Zelle in Stundentafel = 0 (kein NRW-Inject)."""
        import openpyxl
        from data.excel_import import ExcelImporter
        config = default_school_config()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stundentafel"
        ws.append(["Fach", "Jg. 5"])
        ws.append(["Deutsch", ""])  # Leer = 0
        path = tmp_path / "st_missing.xlsx"
        wb.save(str(path))

        importer = ExcelImporter(path, config)
        importer._open()
        st = importer.import_stundentafel(["Deutsch"])
        assert st.get(5, {}).get("Deutsch", 0) == 0

    def test_import_teachers_comma_separated(self, tmp_path: Path):
        """Neues Fächer-Format (kommagetrennt) importiert alle Fächer."""
        import openpyxl
        from data.excel_import import ExcelImporter
        config = default_school_config()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lehrkräfte"
        ws.append(["Name (Nachname, Vorname)", "Kürzel", "Fächer (kommagetrennt)",
                   "Deputat", "Teilzeit", "Sperrzeiten (z.B. Mo1,Di3,Fr5)",
                   "Wunschtage (z.B. Mo,Fr)", "Max Std/Tag", "Max Springstd/Tag"])
        ws.append(["Test, Anna", "TST", "Mathematik, Physik, Informatik, Chemie",
                   26, "nein", "", "", 6, 2])
        path = tmp_path / "lk_komma.xlsx"
        wb.save(str(path))

        importer = ExcelImporter(path, config)
        importer._open()
        teachers = importer.import_teachers()
        assert len(teachers) == 1
        assert len(teachers[0].subjects) == 4
        assert "Informatik" in teachers[0].subjects
        assert "Chemie" in teachers[0].subjects

    def test_import_teachers_old_format_compat(self, tmp_path: Path):
        """Altes Fach 1/2/3-Format wird noch korrekt geparst."""
        import openpyxl
        from data.excel_import import ExcelImporter
        config = default_school_config()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lehrkräfte"
        ws.append(["Name (Nachname, Vorname)", "Kürzel", "Fach 1", "Fach 2", "Fach 3",
                   "Deputat", "Teilzeit", "Sperrzeiten (z.B. Mo1,Di3,Fr5)",
                   "Wunschtage (z.B. Mo,Fr)", "Max Std/Tag", "Max Springstd/Tag"])
        ws.append(["Test, Bob", "BOB", "Mathematik", "Physik", "",
                   26, "nein", "", "", 6, 2])
        path = tmp_path / "lk_old.xlsx"
        wb.save(str(path))

        importer = ExcelImporter(path, config)
        importer._open()
        teachers = importer.import_teachers()
        assert len(teachers) == 1
        assert "Mathematik" in teachers[0].subjects
        assert "Physik" in teachers[0].subjects

    def test_import_classes_uses_custom_stundentafel(self, tmp_path: Path):
        """import_classes() benutzt übergebene Stundentafel statt NRW-Defaults."""
        import openpyxl
        from data.excel_import import ExcelImporter
        config = default_school_config()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jahrgänge"
        ws.append(["Jahrgang", "Anzahl Klassen", "Soll-Stunden/Woche", "Klassen-Buchstaben"])
        ws.append([5, 1, 30, "a"])
        path = tmp_path / "jg.xlsx"
        wb.save(str(path))

        custom_st = {5: {"Deutsch": 5, "Mathematik": 4}}
        importer = ExcelImporter(path, config)
        importer._open()
        classes = importer.import_classes(stundentafel=custom_st)
        assert len(classes) == 1
        assert classes[0].curriculum.get("Deutsch") == 5
        assert classes[0].curriculum.get("Mathematik") == 4

    def test_import_teachers_sperrslots(self, tmp_path: Path):
        """Neues Sperrslots-Format 'Mo:3,Di:1' wird korrekt in unavailable_slots geparsed."""
        import openpyxl
        from data.excel_import import ExcelImporter
        config = default_school_config()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lehrkräfte"
        ws.append([
            "Name (Nachname, Vorname)", "Kürzel", "Fächer (kommagetrennt)",
            "Deputat", "Teilzeit", "Sperrzeiten (z.B. Mo1,Di3,Fr5)",
            "Wunschtage (z.B. Mo,Fr)", "Max Std/Tag", "Max Springstd/Tag",
            "Sperrslots (Tag:Slot,...)", "Wunsch-frei (Tage)", "Max Springstd/Woche",
        ])
        ws.append([
            "Müller, Hans", "TST", "Mathematik",
            26, "nein", "", "", 6, 2,
            "Mo:3,Di:1", "", "",
        ])
        path = tmp_path / "lk.xlsx"
        wb.save(str(path))

        importer = ExcelImporter(path, config)
        importer._open()
        teachers = importer.import_teachers()
        assert len(teachers) == 1
        slots = teachers[0].unavailable_slots
        assert (0, 3) in slots, f"Mo:3 sollte als (0,3) geparsed werden, got {slots}"
        assert (1, 1) in slots, f"Di:1 sollte als (1,1) geparsed werden, got {slots}"

    def test_import_teachers_wunsch_frei(self, tmp_path: Path):
        """Wunsch-frei-Format 'Fr' wird korrekt in preferred_free_days geparsed."""
        import openpyxl
        from data.excel_import import ExcelImporter
        config = default_school_config()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lehrkräfte"
        ws.append([
            "Name (Nachname, Vorname)", "Kürzel", "Fächer (kommagetrennt)",
            "Deputat", "Teilzeit", "Sperrzeiten (z.B. Mo1,Di3,Fr5)",
            "Wunschtage (z.B. Mo,Fr)", "Max Std/Tag", "Max Springstd/Tag",
            "Sperrslots (Tag:Slot,...)", "Wunsch-frei (Tage)", "Max Springstd/Woche",
        ])
        ws.append([
            "Müller, Hans", "TST", "Mathematik",
            26, "nein", "", "", 6, 2,
            "", "Fr Mo", "",
        ])
        path = tmp_path / "lk.xlsx"
        wb.save(str(path))

        importer = ExcelImporter(path, config)
        importer._open()
        teachers = importer.import_teachers()
        assert len(teachers) == 1
        days = teachers[0].preferred_free_days
        assert 4 in days, f"Fr sollte als 4 geparsed werden, got {days}"
        assert 0 in days, f"Mo sollte als 0 geparsed werden, got {days}"

    def test_import_teachers_max_gaps_week(self, tmp_path: Path):
        """Max Springstd/Woche wird pro Lehrer korrekt eingelesen."""
        import openpyxl
        from data.excel_import import ExcelImporter
        config = default_school_config()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lehrkräfte"
        ws.append([
            "Name (Nachname, Vorname)", "Kürzel", "Fächer (kommagetrennt)",
            "Deputat", "Teilzeit", "Sperrzeiten (z.B. Mo1,Di3,Fr5)",
            "Wunschtage (z.B. Mo,Fr)", "Max Std/Tag", "Max Springstd/Tag",
            "Sperrslots (Tag:Slot,...)", "Wunsch-frei (Tage)", "Max Springstd/Woche",
        ])
        ws.append([
            "Müller, Hans", "TST", "Mathematik",
            26, "nein", "", "", 6, 2,
            "", "", 3,
        ])
        path = tmp_path / "lk.xlsx"
        wb.save(str(path))

        importer = ExcelImporter(path, config)
        importer._open()
        teachers = importer.import_teachers()
        assert len(teachers) == 1
        assert teachers[0].max_gaps_per_week == 3


# ─── CSV IMPORT ───────────────────────────────────────────────────────────────

class TestCsvImport:
    def test_single_csv_treated_as_lehrkraefte(self, tmp_path: Path):
        """Einzelne .csv-Datei wird als Lehrkräfte-Blatt behandelt."""
        import csv
        from data.excel_import import CsvImporter
        config = default_school_config()

        csv_path = tmp_path / "Lehrkraefte.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Name (Nachname, Vorname)", "Kürzel",
                             "Fächer (kommagetrennt)", "Deputat", "Teilzeit",
                             "Sperrzeiten (z.B. Mo1,Di3,Fr5)",
                             "Wunschtage (z.B. Mo,Fr)", "Max Std/Tag",
                             "Max Springstd/Tag"])
            writer.writerow(["Csv, Tester", "CSV", "Mathematik, Physik",
                             26, "nein", "", "", 6, 2])

        importer = CsvImporter(csv_path, config)
        importer._open()
        sheet = importer._get_sheet("Lehrkräfte")
        assert sheet is not None
        rows = importer._sheet_rows(sheet)
        assert len(rows) == 1
        assert rows[0]["kürzel"] == "CSV"

    def test_directory_with_multiple_csvs(self, tmp_path: Path):
        """Verzeichnis mit mehreren CSV-Dateien — alle werden erkannt."""
        import csv
        from data.excel_import import CsvImporter
        config = default_school_config()

        csv_dir = tmp_path / "csv_import"
        csv_dir.mkdir()

        # Lehrkräfte
        with open(csv_dir / "Lehrkraefte.csv", "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Name (Nachname, Vorname)", "Kürzel",
                             "Fächer (kommagetrennt)", "Deputat", "Teilzeit",
                             "Sperrzeiten (z.B. Mo1,Di3,Fr5)",
                             "Wunschtage (z.B. Mo,Fr)", "Max Std/Tag",
                             "Max Springstd/Tag"])
            writer.writerow(["Dir, Tester", "DIR", "Mathematik", 26, "nein",
                             "", "", 6, 2])

        importer = CsvImporter(csv_dir, config)
        importer._open()
        assert "Lehrkräfte" in importer._csv_sheets

    def test_missing_csv_uses_defaults_gracefully(self, tmp_path: Path):
        """Fehlende CSV-Dateien lösen keinen Fehler aus (Defaults gelten)."""
        import csv
        from data.excel_import import CsvImporter, ExcelImportError
        config = default_school_config()

        csv_dir = tmp_path / "sparse"
        csv_dir.mkdir()

        # Nur Lehrkräfte vorhanden
        with open(csv_dir / "Lehrkraefte.csv", "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Name (Nachname, Vorname)", "Kürzel",
                             "Fächer (kommagetrennt)", "Deputat", "Teilzeit",
                             "Sperrzeiten (z.B. Mo1,Di3,Fr5)",
                             "Wunschtage (z.B. Mo,Fr)", "Max Std/Tag",
                             "Max Springstd/Tag"])
            writer.writerow(["Sparse, Test", "SPR", "Mathematik", 26, "nein",
                             "", "", 6, 2])

        importer = CsvImporter(csv_dir, config)
        importer._open()
        # Fächer-Blatt fehlt → _get_sheet gibt None → import_subjects fällt zurück
        sheet = importer._get_sheet("Fächer")
        assert sheet is None
        subjects = importer.import_subjects()
        assert len(subjects) == len(SUBJECT_METADATA)


# ─── MAIN.PY CLI ──────────────────────────────────────────────────────────────

class TestCli:
    def test_help(self):
        """main.py --help gibt Usage aus."""
        from click.testing import CliRunner
        from main import cli
        runner = CliRunner()
        result = runner.invoke(cli, ["--help"])
        assert result.exit_code == 0
        assert "Usage" in result.output

    def test_config_show_no_file(self):
        """config show ohne Konfiguration → Fehlermeldung."""
        from click.testing import CliRunner
        from main import cli
        runner = CliRunner()
        with runner.isolated_filesystem():
            result = runner.invoke(cli, ["config", "show"])
            assert result.exit_code != 0 or "Keine Konfiguration" in result.output

    def test_generate_command_exists(self):
        """generate Befehl ist registriert und hat --export-json."""
        from click.testing import CliRunner
        from main import cli
        runner = CliRunner()
        result = runner.invoke(cli, ["generate", "--help"])
        assert result.exit_code == 0
        assert "export-json" in result.output

    def test_template_command_exists(self):
        """template Befehl ist registriert."""
        from click.testing import CliRunner
        from main import cli
        runner = CliRunner()
        result = runner.invoke(cli, ["template", "--help"])
        assert result.exit_code == 0

    def test_import_command_exists(self):
        """import Befehl ist registriert."""
        from click.testing import CliRunner
        from main import cli
        runner = CliRunner()
        result = runner.invoke(cli, ["import", "--help"])
        assert result.exit_code == 0

    def test_validate_command_exists(self):
        """validate Befehl ist registriert."""
        from click.testing import CliRunner
        from main import cli
        runner = CliRunner()
        result = runner.invoke(cli, ["validate", "--help"])
        assert result.exit_code == 0

    def test_solve_command_exists(self):
        """solve Befehl ist registriert."""
        from click.testing import CliRunner
        from main import cli
        runner = CliRunner()
        result = runner.invoke(cli, ["solve", "--help"])
        assert result.exit_code == 0

    def test_scenario_list_command_exists(self):
        """scenario list Befehl ist registriert."""
        from click.testing import CliRunner
        from main import cli
        runner = CliRunner()
        result = runner.invoke(cli, ["scenario", "list"])
        assert result.exit_code == 0


class TestUntisImport:
    """Untis XML Import."""

    def _make_minimal_xml(self) -> str:
        return """<?xml version="1.0" encoding="UTF-8"?>
<untis>
  <subjects>
    <subject id="1"><shortname>D</shortname><longname>Deutsch</longname></subject>
    <subject id="2"><shortname>M</shortname><longname>Mathematik</longname></subject>
    <subject id="99"><shortname>XY</shortname><longname>UnbekanntXY123</longname></subject>
  </subjects>
  <teachers>
    <teacher id="1"><shortname>MUE</shortname><surname>Müller</surname>
      <firstname>Anna</firstname><subjects>Deutsch,Mathematik</subjects></teacher>
    <teacher id="2"><shortname>SCH</shortname><surname>Schmidt</surname>
      <subjects>Mathematik</subjects></teacher>
  </teachers>
  <classes>
    <class id="5a"><shortname>5a</shortname><name>5a</name><grade>5</grade></class>
  </classes>
  <lessons>
    <lesson id="1">
      <teacher id="1"/>
      <subject id="1"/>
      <class id="5a"/>
      <periods>
        <period day="1" period="3"/>
      </periods>
    </lesson>
    <lesson id="2">
      <teacher id="2"/>
      <subject id="2"/>
      <class id="5a"/>
      <periods>
        <period day="2" period="1"/>
        <period day="4" period="1"/>
      </periods>
    </lesson>
  </lessons>
</untis>"""

    def test_parse_subjects(self, tmp_path):
        from data.untis_import import UntisXmlImporter
        xml_file = tmp_path / "test.xml"
        xml_file.write_text(self._make_minimal_xml(), encoding="utf-8")
        config = default_school_config()
        importer = UntisXmlImporter(xml_file, config)
        subjects = importer.import_subjects()
        names = [s.name for s in subjects]
        assert "Deutsch" in names
        assert "Mathematik" in names

    def test_unknown_subject_warning(self, tmp_path):
        from data.untis_import import UntisXmlImporter
        xml_file = tmp_path / "test.xml"
        xml_file.write_text(self._make_minimal_xml(), encoding="utf-8")
        config = default_school_config()
        importer = UntisXmlImporter(xml_file, config)
        importer.import_subjects()
        assert any(
            "UnbekanntXY123" in w or "unbekannt" in w.lower()
            for w in importer._report.warnings
        ), f"Erwartete Warnung, stattdessen: {importer._report.warnings}"

    def test_missing_rooms_section(self, tmp_path):
        from data.untis_import import UntisXmlImporter
        xml_file = tmp_path / "test.xml"
        xml_file.write_text(self._make_minimal_xml(), encoding="utf-8")
        config = default_school_config()
        importer = UntisXmlImporter(xml_file, config)
        rooms = importer.import_rooms()
        assert rooms == []
        assert any(
            "rooms" in w.lower() or "räume" in w.lower() or "raum" in w.lower()
            for w in importer._report.warnings
        )

    def test_parse_teachers(self, tmp_path):
        from data.untis_import import UntisXmlImporter
        xml_file = tmp_path / "test.xml"
        xml_file.write_text(self._make_minimal_xml(), encoding="utf-8")
        config = default_school_config()
        importer = UntisXmlImporter(xml_file, config)
        teachers = importer.import_teachers()
        assert len(teachers) == 2
        ids = [t.id for t in teachers]
        assert "MUE" in ids

    def test_import_lessons_basic(self, tmp_path):
        """2 Lektionen → 3 PinnedLesson-Objekte (Lektion 2 hat 2 Perioden)."""
        from data.untis_import import UntisXmlImporter
        xml_file = tmp_path / "test.xml"
        xml_file.write_text(self._make_minimal_xml(), encoding="utf-8")
        config = default_school_config()
        importer = UntisXmlImporter(xml_file, config)
        teachers = importer.import_teachers()
        classes = importer.import_classes({5: {"Deutsch": 4, "Mathematik": 4}})
        subjects = importer.import_subjects()
        pins = importer.import_lessons(teachers, classes, subjects)
        # Lektion 1: 1 Periode; Lektion 2: 2 Perioden → 3 Pins
        assert len(pins) == 3
        days = [p.day for p in pins]
        assert 0 in days   # Lektion 1: Untis day=1 → solver day=0
        assert 1 in days   # Lektion 2: Untis day=2 → solver day=1

    def test_import_lessons_multi_period(self, tmp_path):
        """Lektion mit 2 Perioden erzeugt 2 PinnedLesson-Objekte."""
        from data.untis_import import UntisXmlImporter
        xml_file = tmp_path / "test.xml"
        xml_file.write_text(self._make_minimal_xml(), encoding="utf-8")
        config = default_school_config()
        importer = UntisXmlImporter(xml_file, config)
        teachers = importer.import_teachers()
        classes = importer.import_classes({5: {"Deutsch": 4, "Mathematik": 4}})
        subjects = importer.import_subjects()
        pins = importer.import_lessons(teachers, classes, subjects)
        # Pins für Lektion 2 (Mathematik, SCH): day=1 slot=1 und day=3 slot=1
        math_pins = [p for p in pins if p.subject == "Mathematik"]
        assert len(math_pins) == 2
        assert any(p.day == 1 and p.slot_number == 1 for p in math_pins)
        assert any(p.day == 3 and p.slot_number == 1 for p in math_pins)

    def test_import_lessons_unknown_ref_warns(self, tmp_path):
        """Unbekannte Lehrer-Referenz erzeugt Warnung und Pin wird übersprungen."""
        from data.untis_import import UntisXmlImporter
        xml = """<?xml version="1.0" encoding="UTF-8"?>
<untis>
  <subjects>
    <subject id="1"><shortname>D</shortname><longname>Deutsch</longname></subject>
  </subjects>
  <teachers>
    <teacher id="1"><shortname>MUE</shortname><subjects>Deutsch</subjects></teacher>
  </teachers>
  <classes>
    <class id="5a"><shortname>5a</shortname><name>5a</name><grade>5</grade></class>
  </classes>
  <lessons>
    <lesson id="99">
      <teacher id="999"/>
      <subject id="1"/>
      <class id="5a"/>
      <periods><period day="1" period="1"/></periods>
    </lesson>
  </lessons>
</untis>"""
        xml_file = tmp_path / "warn.xml"
        xml_file.write_text(xml, encoding="utf-8")
        config = default_school_config()
        importer = UntisXmlImporter(xml_file, config)
        teachers = importer.import_teachers()
        classes = importer.import_classes({5: {"Deutsch": 4}})
        subjects = importer.import_subjects()
        pins = importer.import_lessons(teachers, classes, subjects)
        assert pins == [], f"Unbekannte Referenz sollte keine Pins erzeugen, got {pins}"
        assert any("999" in w or "übersprungen" in w.lower()
                   for w in importer._report.warnings), (
            f"Erwartete Warnung für unbekannte Lehrer-Referenz: {importer._report.warnings}"
        )


class TestDataVersioning:
    """Zeitstempel-Versionierung für SchoolData."""

    def test_save_json_sets_timestamps(self, tmp_path):
        from models.school_data import SchoolData
        data = SchoolData(
            subjects=[], rooms=[], classes=[], teachers=[], couplings=[],
            config=default_school_config(),
        )
        path = tmp_path / "school_data.json"
        data.save_json(path)
        loaded = SchoolData.load_json(path)
        assert loaded.created_at is not None
        assert loaded.modified_at is not None

    def test_save_versioned_filename(self, tmp_path):
        from models.school_data import SchoolData
        data = SchoolData(
            subjects=[], rooms=[], classes=[], teachers=[], couplings=[],
            config=default_school_config(),
        )
        base = tmp_path / "school_data.json"
        actual = data.save_versioned(base)
        assert actual != base
        assert actual.exists()
        assert "school_data_" in actual.name

    def test_modified_at_updates_on_resave(self, tmp_path):
        import time
        from models.school_data import SchoolData
        data = SchoolData(
            subjects=[], rooms=[], classes=[], teachers=[], couplings=[],
            config=default_school_config(),
        )
        path = tmp_path / "school_data.json"
        data.save_json(path)
        loaded1 = SchoolData.load_json(path)

        time.sleep(0.01)
        loaded1.save_json(path)
        loaded2 = SchoolData.load_json(path)

        assert loaded2.modified_at is not None
        assert loaded1.created_at == loaded2.created_at
