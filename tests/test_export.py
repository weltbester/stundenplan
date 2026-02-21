"""Tests für Phase 4: Raum-Zuweisung und Export (Excel + PDF)."""

import pytest
from pathlib import Path

from config.schema import (
    SchoolConfig, GradeConfig, GradeDefinition, SchoolType,
    TeacherConfig, SolverConfig,
)
from config.defaults import default_time_grid, default_rooms, SUBJECT_METADATA, STUNDENTAFEL_GYMNASIUM_SEK1
from models.school_data import SchoolData
from models.subject import Subject
from models.room import Room
from models.teacher import Teacher
from models.school_class import SchoolClass
from models.coupling import Coupling, CouplingGroup
from solver.scheduler import ScheduleSolver, ScheduleSolution, ScheduleEntry
from export.helpers import (
    count_gaps, detect_double_starts, get_coupling_label, build_time_grid_rows,
)
from export.excel_export import ExcelExporter
from export.pdf_export import PdfExporter


# ─── Testdaten-Hilfsfunktionen ────────────────────────────────────────────────

def _make_mini_config() -> SchoolConfig:
    return SchoolConfig(
        school_name="Export-Test-Gymnasium",
        school_type=SchoolType.GYMNASIUM,
        bundesland="NRW",
        time_grid=default_time_grid(),
        grades=GradeConfig(grades=[
            GradeDefinition(grade=5, num_classes=1, weekly_hours_target=30),
            GradeDefinition(grade=7, num_classes=1, weekly_hours_target=32),
        ]),
        rooms=default_rooms(),
        teachers=TeacherConfig(
            total_count=10, vollzeit_deputat=26,
            teilzeit_percentage=0.0, deputat_min_fraction=0.80,
        ),
        solver=SolverConfig(
            time_limit_seconds=60,
            num_workers=4,
            weight_deputat_deviation=0,  # Deaktiviert für schnelle Tests
        ),
    )


def _make_mini_school_data() -> SchoolData:
    config = _make_mini_config()
    sek1_max = config.time_grid.sek1_max_slot

    subjects = [
        Subject(
            name=name, short_name=meta["short"], category=meta["category"],
            is_hauptfach=meta["is_hauptfach"], requires_special_room=meta["room"],
            double_lesson_required=meta["double_required"],
            double_lesson_preferred=meta["double_preferred"],
        )
        for name, meta in SUBJECT_METADATA.items()
    ]
    rooms = []
    for rd in config.rooms.special_rooms:
        prefix = rd.room_type[:2].upper()
        for i in range(1, rd.count + 1):
            rooms.append(Room(id=f"{prefix}{i}", room_type=rd.room_type,
                              name=f"{rd.display_name} {i}"))

    classes = [
        SchoolClass(id="5a", grade=5, label="a",
                    curriculum={s: h for s, h in STUNDENTAFEL_GYMNASIUM_SEK1[5].items() if h > 0},
                    max_slot=sek1_max),
        SchoolClass(id="7a", grade=7, label="a",
                    curriculum={s: h for s, h in STUNDENTAFEL_GYMNASIUM_SEK1[7].items() if h > 0},
                    max_slot=sek1_max),
    ]
    dep_max = 9  # 10×9h=90h >> 62h need → Solver-Spielraum
    dep_min = 4  # T08/T09 (Kopplung-only) bekommen max 4h Kopplungsstunden → dep_min ≤ 4
    teachers = [
        Teacher(id="T01", name="Müller, Anna",   subjects=["Deutsch", "Geschichte"],  deputat_max=dep_max, deputat_min=dep_min, max_hours_per_day=6, max_gaps_per_day=2),
        Teacher(id="T02", name="Schmidt, Hans",  subjects=["Mathematik", "Physik"],   deputat_max=dep_max, deputat_min=dep_min, max_hours_per_day=6, max_gaps_per_day=2),
        Teacher(id="T03", name="Weber, Eva",     subjects=["Englisch", "Politik"],    deputat_max=dep_max, deputat_min=dep_min, max_hours_per_day=6, max_gaps_per_day=2),
        Teacher(id="T04", name="Becker, Klaus",  subjects=["Biologie", "Erdkunde"],   deputat_max=dep_max, deputat_min=dep_min, max_hours_per_day=6, max_gaps_per_day=2),
        Teacher(id="T05", name="Koch, Lisa",     subjects=["Kunst", "Musik"],         deputat_max=dep_max, deputat_min=dep_min, max_hours_per_day=6, max_gaps_per_day=2),
        Teacher(id="T06", name="Wagner, Tom",    subjects=["Sport", "Chemie"],        deputat_max=dep_max, deputat_min=dep_min, max_hours_per_day=6, max_gaps_per_day=2),
        Teacher(id="T07", name="Braun, Sara",    subjects=["Latein", "Deutsch"],      deputat_max=dep_max, deputat_min=dep_min, max_hours_per_day=6, max_gaps_per_day=2),
        Teacher(id="T08", name="Wolf, Peter",    subjects=["Religion", "Ethik"],      deputat_max=dep_max, deputat_min=dep_min, max_hours_per_day=6, max_gaps_per_day=2),
        Teacher(id="T09", name="Neumann, Maria", subjects=["Religion", "Ethik"],      deputat_max=dep_max, deputat_min=dep_min, max_hours_per_day=6, max_gaps_per_day=2),
        Teacher(id="T10", name="Schulz, Ralf",   subjects=["Mathematik", "Deutsch"],  deputat_max=dep_max, deputat_min=dep_min, max_hours_per_day=6, max_gaps_per_day=2),
    ]
    couplings = [
        Coupling(id="reli_5", coupling_type="reli_ethik", involved_class_ids=["5a"],
                 groups=[CouplingGroup(group_name="evangelisch", subject="Religion", hours_per_week=2),
                         CouplingGroup(group_name="ethik",       subject="Ethik",    hours_per_week=2)],
                 hours_per_week=2, cross_class=True),
        Coupling(id="reli_7", coupling_type="reli_ethik", involved_class_ids=["7a"],
                 groups=[CouplingGroup(group_name="evangelisch", subject="Religion", hours_per_week=2),
                         CouplingGroup(group_name="ethik",       subject="Ethik",    hours_per_week=2)],
                 hours_per_week=2, cross_class=True),
    ]
    return SchoolData(subjects=subjects, rooms=rooms, classes=classes,
                      teachers=teachers, couplings=couplings, config=config)


# ─── Fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def mini_school_data() -> SchoolData:
    return _make_mini_school_data()


@pytest.fixture(scope="module")
def mini_solution(mini_school_data: SchoolData) -> ScheduleSolution:
    """Löst den Mini-Datensatz einmal für alle Tests dieses Moduls."""
    solver = ScheduleSolver(mini_school_data)
    solution = solver.solve(use_soft=False)
    assert solution.solver_status in ("OPTIMAL", "FEASIBLE"), (
        f"Solver konnte keine Lösung finden: {solution.solver_status}"
    )
    return solution


# ─── Tests: Raum-Zuweisung ────────────────────────────────────────────────────

class TestRoomAssignment:

    def test_assigns_concrete_ids(self, mini_solution: ScheduleSolution, mini_school_data: SchoolData):
        """Kein Entry enthält noch einen room_type-String; alle room-Werte sind echte Room.id."""
        valid_room_ids = {r.id for r in mini_school_data.rooms}
        room_types = {r.room_type for r in mini_school_data.rooms}

        for entry in mini_solution.entries:
            if entry.room is not None:
                # room darf kein room_type-String mehr sein
                assert entry.room not in room_types, (
                    f"Entry hat noch room_type '{entry.room}' statt konkreter ID: {entry}"
                )
                # room muss eine bekannte Room.id sein (oder "-?" Fallback)
                assert entry.room in valid_room_ids or entry.room.endswith("-?"), (
                    f"Unbekannte room-ID: '{entry.room}'"
                )

    def test_no_double_booking(self, mini_solution: ScheduleSolution):
        """Kein Raum ist an einem (day, slot) doppelt belegt."""
        from collections import defaultdict
        usage: dict[tuple, list[str]] = defaultdict(list)
        for entry in mini_solution.entries:
            if entry.room and not entry.room.endswith("-?"):
                usage[(entry.day, entry.slot_number, entry.room)].append(entry.class_id)

        for (day, slot, room_id), classes in usage.items():
            assert len(classes) == 1, (
                f"Raum {room_id} am Tag {day} Slot {slot} doppelt belegt: {classes}"
            )


# ─── Tests: Helpers ───────────────────────────────────────────────────────────

class TestCountGaps:

    def test_no_gaps(self):
        """Keine Springstunden wenn Stunden konsekutiv."""
        entries = [
            ScheduleEntry(day=0, slot_number=1, teacher_id="T01", class_id="5a", subject="De"),
            ScheduleEntry(day=0, slot_number=2, teacher_id="T01", class_id="5a", subject="De"),
            ScheduleEntry(day=0, slot_number=3, teacher_id="T01", class_id="5a", subject="Ma"),
        ]
        assert count_gaps(entries) == 0

    def test_one_gap(self):
        """Eine Springstunde: Slots 1, 3 (Lücke bei 2)."""
        entries = [
            ScheduleEntry(day=0, slot_number=1, teacher_id="T01", class_id="5a", subject="De"),
            ScheduleEntry(day=0, slot_number=3, teacher_id="T01", class_id="5a", subject="Ma"),
        ]
        assert count_gaps(entries) == 1

    def test_two_gaps(self):
        """Zwei Springstunden: Slots 1, 4 (Lücken bei 2, 3)."""
        entries = [
            ScheduleEntry(day=0, slot_number=1, teacher_id="T01", class_id="5a", subject="De"),
            ScheduleEntry(day=0, slot_number=4, teacher_id="T01", class_id="5a", subject="Ma"),
        ]
        assert count_gaps(entries) == 2

    def test_multiple_days(self):
        """Springstunden über mehrere Tage korrekt summiert."""
        entries = [
            # Tag 0: Slots 1, 3 → 1 Springstunde
            ScheduleEntry(day=0, slot_number=1, teacher_id="T01", class_id="5a", subject="De"),
            ScheduleEntry(day=0, slot_number=3, teacher_id="T01", class_id="5a", subject="Ma"),
            # Tag 1: Slots 2, 5 → 2 Springstunden
            ScheduleEntry(day=1, slot_number=2, teacher_id="T01", class_id="5a", subject="En"),
            ScheduleEntry(day=1, slot_number=5, teacher_id="T01", class_id="5a", subject="Ge"),
        ]
        assert count_gaps(entries) == 3

    def test_empty(self):
        assert count_gaps([]) == 0

    def test_single_entry(self):
        """Ein einzelner Eintrag → keine Springstunden."""
        entries = [ScheduleEntry(day=0, slot_number=3, teacher_id="T01", class_id="5a", subject="De")]
        assert count_gaps(entries) == 0


class TestDetectDoubleStarts:

    def _make_double_blocks(self):
        from config.schema import DoubleBlock
        return [
            DoubleBlock(slot_first=1, slot_second=2),
            DoubleBlock(slot_first=3, slot_second=4),
        ]

    def test_detects_double(self):
        """Doppelstunde wird erkannt wenn beide Slots gleiche Kombination haben."""
        entries = [
            ScheduleEntry(day=0, slot_number=1, teacher_id="T01", class_id="5a", subject="Physik"),
            ScheduleEntry(day=0, slot_number=2, teacher_id="T01", class_id="5a", subject="Physik"),
        ]
        result = detect_double_starts(entries, self._make_double_blocks())
        assert (0, 1) in result

    def test_no_double_different_subject(self):
        """Keine Doppelstunde wenn Fächer unterschiedlich."""
        entries = [
            ScheduleEntry(day=0, slot_number=1, teacher_id="T01", class_id="5a", subject="Physik"),
            ScheduleEntry(day=0, slot_number=2, teacher_id="T01", class_id="5a", subject="Chemie"),
        ]
        result = detect_double_starts(entries, self._make_double_blocks())
        assert len(result) == 0

    def test_no_double_different_day(self):
        """Keine Doppelstunde wenn verschiedene Tage."""
        entries = [
            ScheduleEntry(day=0, slot_number=1, teacher_id="T01", class_id="5a", subject="Physik"),
            ScheduleEntry(day=1, slot_number=2, teacher_id="T01", class_id="5a", subject="Physik"),
        ]
        result = detect_double_starts(entries, self._make_double_blocks())
        assert len(result) == 0

    def test_multiple_doubles(self):
        """Zwei Doppelstunden an verschiedenen Tagen erkannt."""
        entries = [
            ScheduleEntry(day=0, slot_number=1, teacher_id="T01", class_id="5a", subject="Physik"),
            ScheduleEntry(day=0, slot_number=2, teacher_id="T01", class_id="5a", subject="Physik"),
            ScheduleEntry(day=2, slot_number=3, teacher_id="T02", class_id="7a", subject="Chemie"),
            ScheduleEntry(day=2, slot_number=4, teacher_id="T02", class_id="7a", subject="Chemie"),
        ]
        result = detect_double_starts(entries, self._make_double_blocks())
        assert (0, 1) in result
        assert (2, 3) in result


class TestGetCouplingLabel:

    def test_returns_group_name_for_reli(self, mini_school_data: SchoolData):
        """Gibt group_name für reli_ethik-Kopplung zurück."""
        entry = ScheduleEntry(
            day=0, slot_number=1, teacher_id="T08", class_id="5a",
            subject="Religion", is_coupling=True, coupling_id="reli_5",
        )
        label = get_coupling_label(entry, mini_school_data)
        assert label == "evangelisch"

    def test_returns_group_name_for_ethik(self, mini_school_data: SchoolData):
        entry = ScheduleEntry(
            day=0, slot_number=2, teacher_id="T09", class_id="5a",
            subject="Ethik", is_coupling=True, coupling_id="reli_5",
        )
        label = get_coupling_label(entry, mini_school_data)
        assert label == "ethik"

    def test_returns_none_for_non_coupling(self, mini_school_data: SchoolData):
        entry = ScheduleEntry(
            day=0, slot_number=1, teacher_id="T01", class_id="5a", subject="Deutsch",
        )
        assert get_coupling_label(entry, mini_school_data) is None

    def test_returns_none_for_no_coupling_id(self, mini_school_data: SchoolData):
        entry = ScheduleEntry(
            day=0, slot_number=1, teacher_id="T01", class_id="5a",
            subject="Religion", is_coupling=True,  # coupling_id fehlt
        )
        assert get_coupling_label(entry, mini_school_data) is None


# ─── Tests: Excel-Export ──────────────────────────────────────────────────────

class TestExcelExport:

    def test_creates_file(self, tmp_path: Path, mini_solution: ScheduleSolution, mini_school_data: SchoolData):
        """ExcelExporter.export() erzeugt eine .xlsx-Datei."""
        out = tmp_path / "test_stundenplan.xlsx"
        ExcelExporter(mini_solution, mini_school_data).export(out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_has_correct_sheets(self, tmp_path: Path, mini_solution: ScheduleSolution, mini_school_data: SchoolData):
        """Pflicht-Sheets vorhanden: Übersicht, Klasse 5a, Klasse 7a, Lehrer T01, Raum (mind. 1)."""
        from openpyxl import load_workbook
        out = tmp_path / "test_sheets.xlsx"
        ExcelExporter(mini_solution, mini_school_data).export(out)
        wb = load_workbook(out)
        titles = [ws.title for ws in wb.worksheets]

        assert "Übersicht" in titles
        assert "Klasse 5a" in titles
        assert "Klasse 7a" in titles
        assert "Lehrer T01" in titles
        # Mindestens ein Raum-Sheet muss vorhanden sein (Fachräume werden belegt)
        raum_sheets = [t for t in titles if t.startswith("Raum ")]
        assert len(raum_sheets) >= 1, f"Kein Raum-Sheet gefunden in: {titles}"

    def test_uebersicht_has_teacher_names(self, tmp_path: Path, mini_solution: ScheduleSolution, mini_school_data: SchoolData):
        """Übersicht-Sheet enthält Lehrer-Namen."""
        from openpyxl import load_workbook
        out = tmp_path / "test_uebersicht.xlsx"
        ExcelExporter(mini_solution, mini_school_data).export(out)
        wb = load_workbook(out)
        ws = wb["Übersicht"]
        all_values = {cell.value for row in ws.iter_rows() for cell in row if cell.value}
        # Mindestens einer der Lehrer-Namen sollte im Sheet stehen
        assert any("Müller" in str(v) or "T01" in str(v) for v in all_values)

    def test_class_sheet_has_content(self, tmp_path: Path, mini_solution: ScheduleSolution, mini_school_data: SchoolData):
        """Klassen-Sheet enthält Fach-Einträge."""
        from openpyxl import load_workbook
        out = tmp_path / "test_class.xlsx"
        ExcelExporter(mini_solution, mini_school_data).export(out)
        wb = load_workbook(out)
        ws = wb["Klasse 5a"]
        # Mindestens eine Zelle mit einem bekannten Fach-Namen
        all_values = {str(cell.value) for row in ws.iter_rows() for cell in row if cell.value}
        subjects_in_sheet = {v for v in all_values if any(
            subj in v for subj in ["Deutsch", "Mathematik", "Englisch", "Biologie"]
        )}
        assert len(subjects_in_sheet) >= 1, "Kein Fach-Eintrag im Klassen-Sheet"


# ─── Tests: PDF-Export ────────────────────────────────────────────────────────

class TestPdfExport:

    def test_creates_class_pdf(self, tmp_path: Path, mini_solution: ScheduleSolution, mini_school_data: SchoolData):
        """export_class_schedules() erzeugt eine nicht-leere PDF."""
        out = tmp_path / "klassen.pdf"
        PdfExporter(mini_solution, mini_school_data).export_class_schedules(out)
        assert out.exists()
        assert out.stat().st_size > 1000   # Mindestgröße für eine echte PDF

    def test_creates_teacher_pdf(self, tmp_path: Path, mini_solution: ScheduleSolution, mini_school_data: SchoolData):
        """export_teacher_schedules() erzeugt eine nicht-leere PDF."""
        out = tmp_path / "lehrer.pdf"
        PdfExporter(mini_solution, mini_school_data).export_teacher_schedules(out)
        assert out.exists()
        assert out.stat().st_size > 1000

    def test_both_pdfs_are_valid(self, tmp_path: Path, mini_solution: ScheduleSolution, mini_school_data: SchoolData):
        """Beide PDFs beginnen mit dem PDF-Magic-Byte '%PDF'."""
        classes_out  = tmp_path / "k.pdf"
        teachers_out = tmp_path / "l.pdf"
        exp = PdfExporter(mini_solution, mini_school_data)
        exp.export_class_schedules(classes_out)
        exp.export_teacher_schedules(teachers_out)

        for pdf_file in (classes_out, teachers_out):
            header = pdf_file.read_bytes()[:4]
            assert header == b"%PDF", f"{pdf_file.name} ist keine gültige PDF"


# ─── Tests: build_time_grid_rows ─────────────────────────────────────────────

class TestBuildTimeGridRows:

    def test_row_count(self, mini_school_data: SchoolData):
        """Korrekte Anzahl von Zeilen (Sek-I-Slots + Pausen)."""
        rows = build_time_grid_rows(mini_school_data.config)
        from config.schema import LessonSlot, PauseSlot
        lesson_rows = [r for r in rows if isinstance(r, LessonSlot)]
        pause_rows  = [r for r in rows if isinstance(r, PauseSlot)]
        tg = mini_school_data.config.time_grid
        expected_lessons = len([s for s in tg.lesson_slots if not s.is_sek2_only and s.slot_number <= tg.sek1_max_slot])
        assert len(lesson_rows) == expected_lessons

    def test_pause_after_slot(self, mini_school_data: SchoolData):
        """Pausen folgen auf den richtigen Slot."""
        from config.schema import LessonSlot, PauseSlot
        rows = build_time_grid_rows(mini_school_data.config)
        for i, row in enumerate(rows):
            if isinstance(row, PauseSlot):
                assert i > 0, "Pause als erste Zeile ist ungültig"
                prev = rows[i - 1]
                assert isinstance(prev, LessonSlot)
                assert prev.slot_number == row.after_slot
