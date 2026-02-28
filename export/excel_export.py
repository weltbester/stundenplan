"""Excel-Export für den Stundenplan (openpyxl)."""

from collections import defaultdict
from pathlib import Path
from typing import Optional

from config.schema import LessonSlot, PauseSlot
from models.school_data import SchoolData
from models.teacher import Teacher
from models.room import Room
from solver.scheduler import ScheduleEntry, ScheduleSolution

from export.helpers import (
    COLORS, build_time_grid_rows, get_subject_color, count_gaps,
    detect_double_starts, count_teacher_actual_hours,
    format_entries, today_str,
)


class ExcelExporter:
    """Exportiert eine ScheduleSolution in eine Excel-Datei mit 4 Sheet-Typen."""

    # Spaltenbreiten (Excel-Einheiten)
    COL_STD_W  = 6
    COL_ZEIT_W = 15
    COL_DAY_W  = 22

    # Zeilenhöhen (Punkte)
    ROW_HEADER_H  = 22
    ROW_LESSON_H  = 48
    ROW_PAUSE_H   = 12

    def __init__(self, solution: ScheduleSolution, school_data: SchoolData):
        self.solution   = solution
        self.data       = school_data
        self.config     = school_data.config
        self.tg         = school_data.config.time_grid
        self.days       = list(range(self.tg.days_per_week))
        self.day_names  = self.tg.day_names

    # ─── Öffentliche API ──────────────────────────────────────────────────────

    def export(self, output_path: Path, quality_report=None) -> None:
        """Erstellt die Excel-Datei mit allen Sheets.

        quality_report: optionaler ScheduleQualityReport – wenn angegeben,
        wird ein zusätzliches Qualitätsblatt eingefügt.
        """
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)   # Leeres Standard-Sheet entfernen

        self._sheet_uebersicht(wb)

        if quality_report is not None:
            self._sheet_qualitaet(wb, quality_report)

        for cls in sorted(self.data.classes, key=lambda c: c.id):
            self._sheet_klasse(wb, cls.id)

        for teacher in sorted(self.data.teachers, key=lambda t: t.id):
            self._sheet_lehrer(wb, teacher)

        used_rooms = {e.room for e in self.solution.entries if e.room}
        for room in sorted(self.data.rooms, key=lambda r: r.id):
            if room.id in used_rooms:
                self._sheet_raum(wb, room)

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

    # ─── Style-Helpers ────────────────────────────────────────────────────────

    def _fill(self, hex_color: str):
        from openpyxl.styles import PatternFill
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def _center_align(self, wrap: bool = True):
        from openpyxl.styles import Alignment
        return Alignment(wrap_text=wrap, horizontal="center", vertical="center")

    def _thin_border(self):
        from openpyxl.styles import Border, Side
        s = Side(border_style="thin", color="BBBBBB")
        return Border(left=s, right=s, top=s, bottom=s)

    def _setup_sheet(self, ws) -> None:
        """Setzt Spaltenbreiten für ein Tabellenblatt."""
        from openpyxl.utils import get_column_letter
        ws.column_dimensions["A"].width = self.COL_STD_W
        ws.column_dimensions["B"].width = self.COL_ZEIT_W
        for col in range(3, 3 + len(self.days)):
            ws.column_dimensions[get_column_letter(col)].width = self.COL_DAY_W

    def _write_header_row(self, ws) -> None:
        """Schreibt die Kopfzeile (Std. | Zeit | Mo | Di | …)."""
        from openpyxl.styles import Font
        headers = ["Std.", "Zeit"] + self.day_names[: len(self.days)]
        fill = self._fill(COLORS["header"])
        border = self._thin_border()
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=text)
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = self._center_align(wrap=False)
            cell.border = border
        ws.row_dimensions[1].height = self.ROW_HEADER_H

    # ─── Eintragsgitter ───────────────────────────────────────────────────────

    def _build_grid(
        self, entries: list[ScheduleEntry]
    ) -> dict[tuple[int, int], list[ScheduleEntry]]:
        """Baut {(day, slot_number): [entries]} für die übergebenen Entries auf."""
        grid: dict[tuple[int, int], list[ScheduleEntry]] = defaultdict(list)
        for e in entries:
            grid[(e.day, e.slot_number)].append(e)
        return grid

    def _cell_color(self, entries: list[ScheduleEntry]) -> str:
        """Gibt die Hintergrundfarbe für eine Zelle zurück."""
        if not entries:
            return COLORS["free"]
        e = entries[0]
        if e.is_coupling:
            return COLORS["coupling"]
        return get_subject_color(e.subject, self.data.subjects)

    # ─── Zeitraster-Tabelle ───────────────────────────────────────────────────

    def _write_schedule_table(
        self, ws, entries: list[ScheduleEntry], mode: str,
        max_slot: int | None = None,
    ) -> int:
        """Schreibt das Zeitraster mit Inhalten; gibt letzte verwendete Excel-Zeile zurück.

        mode: 'class' | 'teacher' | 'room'
        max_slot begrenzt die angezeigten Zeilen auf die für diese Entität
        relevanten Slots (Sek. I: sek1_max_slot, Sek. II: höher).
        """
        from openpyxl.styles import Font

        effective_max = max_slot if max_slot is not None else self.tg.sek1_max_slot
        time_rows = build_time_grid_rows(self.config, effective_max)
        double_blocks = [
            db for db in self.tg.double_blocks
            if db.slot_second <= effective_max
        ]
        double_pair_map = {db.slot_first: db.slot_second for db in double_blocks}

        double_starts = detect_double_starts(entries, double_blocks)
        grid = self._build_grid(entries)
        border = self._thin_border()

        excel_row = 2   # Zeile 1 = Header
        slot_row_map: dict[int, int] = {}   # slot_number → Excel-Zeile

        for row_obj in time_rows:
            if isinstance(row_obj, LessonSlot):
                slot_num = row_obj.slot_number
                slot_row_map[slot_num] = excel_row
                time_str = f"{row_obj.start_time}–{row_obj.end_time}"

                # Std.-Spalte
                c = ws.cell(row=excel_row, column=1, value=slot_num)
                c.alignment = self._center_align(wrap=False)
                c.border = border
                c.font = Font(bold=True, size=9)

                # Zeit-Spalte
                c = ws.cell(row=excel_row, column=2, value=time_str)
                c.alignment = self._center_align(wrap=False)
                c.border = border
                c.font = Font(size=8)

                # Tag-Spalten
                for day in self.days:
                    col = day + 3
                    here = grid.get((day, slot_num), [])
                    content = format_entries(here, self.data, mode)
                    color = self._cell_color(here)

                    # Springstunden für Lehrerplan: rot hinterlegen
                    if mode == "teacher" and not here:
                        teacher_slots_today = [
                            sn for (d, sn), es in grid.items()
                            if d == day and es
                        ]
                        if teacher_slots_today:
                            mn, mx = min(teacher_slots_today), max(teacher_slots_today)
                            if mn < slot_num < mx:
                                color = COLORS["gap"]

                    c = ws.cell(row=excel_row, column=col, value=content)
                    c.fill = self._fill(color)
                    c.alignment = self._center_align()
                    c.border = border
                    c.font = Font(size=8)

                ws.row_dimensions[excel_row].height = self.ROW_LESSON_H

            elif isinstance(row_obj, PauseSlot):
                label = f"── {row_obj.label} ({row_obj.duration_minutes} Min.) ──"
                num_cols = 2 + len(self.days)
                ws.merge_cells(
                    start_row=excel_row, start_column=1,
                    end_row=excel_row, end_column=num_cols,
                )
                c = ws.cell(row=excel_row, column=1, value=label)
                c.fill = self._fill(COLORS["pause"])
                c.alignment = self._center_align(wrap=False)
                c.font = Font(italic=True, size=8, color="666666")
                ws.row_dimensions[excel_row].height = self.ROW_PAUSE_H

            excel_row += 1

        # Doppelstunden-Zellen zusammenführen
        for (day, slot_first) in double_starts:
            slot_second = double_pair_map.get(slot_first)
            if slot_second is None:
                continue
            row1 = slot_row_map.get(slot_first)
            row2 = slot_row_map.get(slot_second)
            if row1 is None or row2 is None:
                continue
            col = day + 3
            # Nur zusammenführen wenn die Zelle nicht leer ist
            if ws.cell(row=row1, column=col).value:
                ws.merge_cells(
                    start_row=row1, start_column=col,
                    end_row=row2, end_column=col,
                )
                # Stil des zusammengeführten Bereichs sicherstellen
                c = ws.cell(row=row1, column=col)
                c.alignment = self._center_align()

        return excel_row

    # ─── Sheet: Übersicht ─────────────────────────────────────────────────────

    def _sheet_uebersicht(self, wb) -> None:
        from openpyxl.styles import Font
        ws = wb.create_sheet(title="Übersicht", index=0)

        row = 1
        # Titel
        ws.cell(row=row, column=1, value=self.config.school_name).font = Font(bold=True, size=14)
        row += 1
        ws.cell(row=row, column=1, value=f"Erstellt: {today_str()}")
        ws.cell(row=row, column=3, value=f"Status: {self.solution.solver_status}")
        ws.cell(row=row, column=4, value=f"Zeit: {self.solution.solve_time_seconds:.1f}s")
        if self.solution.objective_value is not None:
            ws.cell(row=row, column=5, value=f"Obj: {self.solution.objective_value:.0f}")
        row += 2

        # Lehrer-Tabelle
        headers = ["Kürzel", "Name", "Fächer", "Min", "Max", "Ist", "Springstd."]
        fill_h = self._fill(COLORS["header"])
        border = self._thin_border()
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = fill_h
            c.font = Font(bold=True, color="FFFFFF")
            c.border = border
        row += 1

        for teacher in sorted(self.data.teachers, key=lambda t: t.id):
            actual = count_teacher_actual_hours(self.solution.entries, teacher.id)
            t_entries = self.solution.get_teacher_schedule(teacher.id)
            gaps = count_gaps(t_entries)

            ws.cell(row=row, column=1, value=teacher.id).border = border
            ws.cell(row=row, column=2, value=teacher.name).border = border
            ws.cell(row=row, column=3, value=", ".join(teacher.subjects)).border = border
            ws.cell(row=row, column=4, value=teacher.deputat_min).border = border
            ws.cell(row=row, column=5, value=teacher.deputat_max).border = border
            c_ist = ws.cell(row=row, column=6, value=actual)
            c_ist.border = border
            if actual < teacher.deputat_min or actual > teacher.deputat_max:
                c_ist.fill = self._fill("FFCCCC")
            ws.cell(row=row, column=7, value=gaps).border = border
            row += 1

        row += 1

        # Fachraum-Auslastung
        ws.cell(row=row, column=1, value="Fachraum-Auslastung").font = Font(bold=True)
        row += 1
        headers_r = ["Raum-ID", "Name", "Typ", "Belegte Slots", "Max. Slots", "Auslastung"]
        for col, h in enumerate(headers_r, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = fill_h
            c.font = Font(bold=True, color="FFFFFF")
            c.border = border
        row += 1

        max_slots = self.tg.sek1_max_slot * self.tg.days_per_week
        # De-duplizieren nach (room, day, slot) – Kopplungsstunden zählen einmal,
        # auch wenn sie für mehrere Klassen als Eintrag erscheinen.
        seen_room_slots: set[tuple] = set()
        room_usage: dict[str, int] = defaultdict(int)
        for e in self.solution.entries:
            if e.room and not e.room.endswith("-?"):
                key = (e.room, e.day, e.slot_number)
                if key not in seen_room_slots:
                    seen_room_slots.add(key)
                    room_usage[e.room] += 1

        for room in sorted(self.data.rooms, key=lambda r: r.id):
            occupied = room_usage.get(room.id, 0)
            pct = f"{occupied / max_slots * 100:.0f}%" if max_slots else "–"
            ws.cell(row=row, column=1, value=room.id).border = border
            ws.cell(row=row, column=2, value=room.name).border = border
            ws.cell(row=row, column=3, value=room.room_type).border = border
            ws.cell(row=row, column=4, value=occupied).border = border
            ws.cell(row=row, column=5, value=max_slots).border = border
            ws.cell(row=row, column=6, value=pct).border = border
            row += 1

        # Spaltenbreiten
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 12

    # ─── Sheet: Klasse ────────────────────────────────────────────────────────

    def _sheet_klasse(self, wb, class_id: str) -> None:
        title = f"Klasse {class_id}"[:31]
        ws = wb.create_sheet(title=title)
        self._setup_sheet(ws)
        self._write_header_row(ws)
        entries = self.solution.get_class_schedule(class_id)
        cls = next(c for c in self.data.classes if c.id == class_id)
        self._write_schedule_table(ws, entries, mode="class", max_slot=cls.max_slot)

    # ─── Sheet: Lehrer ────────────────────────────────────────────────────────

    def _sheet_lehrer(self, wb, teacher: Teacher) -> None:
        title = f"Lehrer {teacher.id}"[:31]
        ws = wb.create_sheet(title=title)
        self._setup_sheet(ws)
        self._write_header_row(ws)
        entries = self.solution.get_teacher_schedule(teacher.id)
        max_slot = max(
            (e.slot_number for e in entries), default=self.tg.sek1_max_slot
        )
        last_row = self._write_schedule_table(ws, entries, mode="teacher", max_slot=max_slot)

        # Stat-Box unter dem Zeitraster
        from openpyxl.styles import Font
        actual = count_teacher_actual_hours(self.solution.entries, teacher.id)
        gaps = count_gaps(entries)
        last_row += 1
        ws.cell(row=last_row, column=1, value="Deputat:").font = Font(bold=True)
        ws.cell(row=last_row, column=2, value=f"Min: {teacher.deputat_min}h | Max: {teacher.deputat_max}h")
        ws.cell(row=last_row, column=3, value="Ist:").font = Font(bold=True)
        ws.cell(row=last_row, column=4, value=f"{actual}h")
        ws.cell(row=last_row, column=5, value="Springstunden:").font = Font(bold=True)
        ws.cell(row=last_row, column=6, value=str(gaps))

    # ─── Sheet: Raum ──────────────────────────────────────────────────────────

    def _sheet_raum(self, wb, room: Room) -> None:
        title = f"Raum {room.id}"[:31]
        ws = wb.create_sheet(title=title)
        self._setup_sheet(ws)
        self._write_header_row(ws)
        entries = [e for e in self.solution.entries if e.room == room.id]
        max_slot = max(
            (e.slot_number for e in entries), default=self.tg.sek1_max_slot
        )
        self._write_schedule_table(ws, entries, mode="room", max_slot=max_slot)

    # ─── Sheet: Qualität ──────────────────────────────────────────────────────

    def _sheet_qualitaet(self, wb, report) -> None:
        """Erstellt ein Qualitätsblatt mit drei Tabellen.

        Tab 1: Lehrer-Übersicht (Name, Min, Max, Ist, Gaps, Status)
        Tab 2: Klassen-Übersicht (Klasse, Stunden/Tag, Doppelstunden-Rate)
        Tab 3: KPI-Übersicht (Fairness-Index, Gesamt-Gaps, Doppelstunden-Rate)
        """
        from openpyxl.styles import Font

        ws = wb.create_sheet(title="Qualität")
        fill_h = self._fill(COLORS["header"])
        border = self._thin_border()
        row = 1

        # ── KPI-Block ──────────────────────────────────────────────────────
        ws.cell(row=row, column=1, value="Qualitätsbericht").font = Font(bold=True, size=13)
        row += 1
        ws.cell(row=row, column=1, value=f"Erstellt: {today_str()}")
        ws.cell(row=row, column=3, value=f"Status: {report.solver_status}")
        ws.cell(row=row, column=4, value=f"Zeit: {report.solve_time}s")
        row += 2

        kpi_headers = ["KPI", "Wert", "Bewertung"]
        for col, h in enumerate(kpi_headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = fill_h
            c.font = Font(bold=True, color="FFFFFF")
            c.border = border
        row += 1

        kpis = [
            ("Gesamt-Springstunden", str(report.total_gaps),
             "gut" if report.total_gaps < 20 else "mittel" if report.total_gaps < 50 else "hoch"),
            ("Ø Springstunden/Lehrer", f"{report.avg_gaps_per_teacher:.1f}",
             "gut" if report.avg_gaps_per_teacher < 2 else "mittel"),
            ("Deputat-Fairness (Jain)", f"{report.deputat_fairness_index:.4f}",
             "gut" if report.deputat_fairness_index >= 0.95 else "mittel"),
            ("Doppelstunden-Rate", f"{report.double_fulfillment_rate:.1%}",
             "gut" if report.double_fulfillment_rate >= 0.90 else "mittel"),
        ]
        for name, value, rating in kpis:
            ws.cell(row=row, column=1, value=name).border = border
            ws.cell(row=row, column=2, value=value).border = border
            ws.cell(row=row, column=3, value=rating).border = border
            row += 1

        row += 2

        # ── Lehrer-Tabelle ────────────────────────────────────────────────
        ws.cell(row=row, column=1, value="Lehrer-Auslastung").font = Font(bold=True, size=11)
        row += 1
        t_headers = ["ID", "Name", "Min", "Max", "Ist", "Gaps", "Freie Tage", "Status"]
        for col, h in enumerate(t_headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = fill_h
            c.font = Font(bold=True, color="FFFFFF")
            c.border = border
        row += 1

        for m in sorted(report.teacher_metrics, key=lambda x: x.teacher_id):
            if m.actual_hours < m.dep_min:
                status = "Unter Min"
                status_fill = self._fill("FFCCCC")
            elif m.actual_hours > m.dep_max:
                status = "Über Max"
                status_fill = self._fill("FFEECC")
            else:
                status = "OK"
                status_fill = self._fill("CCFFCC")

            ws.cell(row=row, column=1, value=m.teacher_id).border = border
            ws.cell(row=row, column=2, value=m.name).border = border
            ws.cell(row=row, column=3, value=m.dep_min).border = border
            ws.cell(row=row, column=4, value=m.dep_max).border = border
            c_ist = ws.cell(row=row, column=5, value=m.actual_hours)
            c_ist.border = border
            ws.cell(row=row, column=6, value=m.gaps_total).border = border
            ws.cell(row=row, column=7, value=m.free_days).border = border
            c_status = ws.cell(row=row, column=8, value=status)
            c_status.border = border
            c_status.fill = status_fill
            row += 1

        row += 2

        # ── Klassen-Tabelle ───────────────────────────────────────────────
        ws.cell(row=row, column=1, value="Klassen-Qualität").font = Font(bold=True, size=11)
        row += 1
        c_headers = ["Klasse", "Stunden", "Doppelstd. Soll", "Doppelstd. Ist", "Spread-Score"]
        for col, h in enumerate(c_headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = fill_h
            c.font = Font(bold=True, color="FFFFFF")
            c.border = border
        row += 1

        for m in sorted(report.class_metrics, key=lambda x: x.class_id):
            ws.cell(row=row, column=1, value=m.class_id).border = border
            ws.cell(row=row, column=2, value=m.total_hours).border = border
            ws.cell(row=row, column=3, value=m.double_requested).border = border
            ws.cell(row=row, column=4, value=m.double_fulfilled).border = border
            c_spread = ws.cell(row=row, column=5, value=m.subject_spread_score)
            c_spread.border = border
            if m.subject_spread_score >= 0.7:
                c_spread.fill = self._fill("CCFFCC")
            elif m.subject_spread_score >= 0.4:
                c_spread.fill = self._fill("FFFFCC")
            else:
                c_spread.fill = self._fill("FFCCCC")
            row += 1

        # Spaltenbreiten
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 12
