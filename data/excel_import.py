"""Excel-Import und Template-Generator für echte Schuldaten.

Template-Generator: Leere Excel-Vorlage mit vorausgefüllten Blättern.
Import-Funktion:    Excel → SchoolData mit Validierung und FeasibilityReport.
"""

import difflib
from pathlib import Path
from typing import Optional

from config.schema import SchoolConfig
from config.defaults import STUNDENTAFEL_GYMNASIUM_SEK1, SUBJECT_METADATA
from models.teacher import Teacher
from models.school_class import SchoolClass
from models.subject import Subject
from models.room import Room
from models.coupling import Coupling, CouplingGroup
from models.school_data import SchoolData, FeasibilityReport


class ExcelImportError(Exception):
    """Fehler beim Excel-Import."""


# ─── Tages-Mapping ────────────────────────────────────────────────────────────

_DAY_MAP = {"mo": 0, "di": 1, "mi": 2, "do": 3, "fr": 4,
            "montag": 0, "dienstag": 1, "mittwoch": 2, "donnerstag": 3, "freitag": 4}

_DAY_NAMES = ["Mo", "Di", "Mi", "Do", "Fr"]


def _parse_blocked_slots(raw: str) -> list[tuple[int, int]]:
    """Parst Sperrzeiten-String 'Mo1,Di3,Fr5' → [(0,1),(1,3),(4,5)]."""
    if not raw.strip():
        return []
    result = []
    for token in raw.replace(";", ",").split(","):
        token = token.strip().lower()
        if not token:
            continue
        # Trenne Tages-Buchstaben von Slot-Zahl
        for prefix in sorted(_DAY_MAP, key=len, reverse=True):
            if token.startswith(prefix):
                slot_str = token[len(prefix):]
                try:
                    slot = int(slot_str)
                    result.append((_DAY_MAP[prefix], slot))
                except ValueError:
                    pass  # Ignoriere ungültige Tokens
                break
    return result


def _parse_free_days(raw: str) -> list[int]:
    """Parst Wunschtage-String 'Mo,Fr' → [0,4]."""
    if not raw.strip():
        return []
    result = []
    for token in raw.replace(";", ",").split(","):
        token = token.strip().lower()
        if token in _DAY_MAP:
            day = _DAY_MAP[token]
            if day not in result:
                result.append(day)
    return result


def _parse_sperrslots(raw: str) -> list[tuple[int, int]]:
    """Parst neues Sperrzeiten-Format 'Mo:3,Di:1' → [(0,3),(1,1)].

    Unterstützt Komma- und Semikolon-Trennung; Tag und Slot durch Doppelpunkt getrennt.
    """
    if not raw.strip():
        return []
    result = []
    for token in raw.replace(";", ",").split(","):
        token = token.strip()
        if ":" not in token:
            continue
        day_str, _, slot_str = token.partition(":")
        day_str = day_str.strip().lower()
        slot_str = slot_str.strip()
        if day_str in _DAY_MAP:
            try:
                result.append((_DAY_MAP[day_str], int(slot_str)))
            except ValueError:
                pass
    return result


def _parse_free_days_flexible(raw: str) -> list[int]:
    """Parst Wunschtage mit Leerzeichen- oder Komma-Trennung 'Fr Mo' → [4,0]."""
    import re
    if not raw.strip():
        return []
    result = []
    for token in re.split(r'[\s,;]+', raw.strip()):
        token = token.strip().lower()
        if token in _DAY_MAP:
            day = _DAY_MAP[token]
            if day not in result:
                result.append(day)
    return result


def _fuzzy_subject(name: str, known: list[str]) -> Optional[str]:
    """Fuzzy-Matching: Findet das ähnlichste bekannte Fach."""
    matches = difflib.get_close_matches(name, known, n=1, cutoff=0.6)
    return matches[0] if matches else None


# ─── TEMPLATE-GENERATOR ───────────────────────────────────────────────────────

def generate_template(config: SchoolConfig, path: Path) -> None:
    """Erzeugt eine leere Excel-Vorlage mit vorausgefüllten Blättern.

    Blätter:
      - Zeitraster:    Slot-Nr, Start, Ende, SII-only (aus Config)
      - Jahrgänge:     Jahrgang, Klassen, Soll-Stunden (aus Config)
      - Stundentafel:  Jahrgang × Fächer Matrix (aus STUNDENTAFEL)
      - Lehrkräfte:    Eingabe-Vorlage mit Beispielzeile
      - Fachräume:     Raumtyp, Name, Anzahl
      - Kopplungen:    Jahrgang, Typ, Klassen, Gruppen, Stunden
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        raise ImportError("openpyxl nicht installiert. Bitte: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ── Hilfs-Styles ─────────────────────────────────────────────────────────
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2E6DA4")
    alt_fill = PatternFill("solid", fgColor="D6E4F0")
    ex_font = Font(italic=True, color="888888")
    ex_fill = PatternFill("solid", fgColor="F5F5F5")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(cell, width: int = 14):
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border

    def style_data(cell, alt: bool = False):
        cell.fill = alt_fill if alt else PatternFill()
        cell.alignment = Alignment(vertical="center")
        cell.border = border

    def style_example(cell):
        cell.font = ex_font
        cell.fill = ex_fill
        cell.border = border

    def set_col_width(ws, col: int, width: float):
        ws.column_dimensions[get_column_letter(col)].width = width

    def write_row(ws, row: int, values: list, style_fn=style_data, alt: bool = False):
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_fn(cell) if style_fn != style_data else style_fn(cell, alt=alt)

    # ── Blatt 1: Zeitraster ───────────────────────────────────────────────────
    ws_zt = wb.active
    ws_zt.title = "Zeitraster"
    headers = ["Slot-Nr", "Beginn", "Ende", "SII-only"]
    for col, h in enumerate(headers, 1):
        cell = ws_zt.cell(row=1, column=col, value=h)
        style_header(cell)
    set_col_width(ws_zt, 1, 10)
    set_col_width(ws_zt, 2, 10)
    set_col_width(ws_zt, 3, 10)
    set_col_width(ws_zt, 4, 12)

    for r, slot in enumerate(config.time_grid.lesson_slots, 2):
        row_vals = [slot.slot_number, slot.start_time, slot.end_time,
                    "ja" if slot.is_sek2_only else "nein"]
        alt = (r % 2 == 0)
        for col, val in enumerate(row_vals, 1):
            cell = ws_zt.cell(row=r, column=col, value=val)
            style_data(cell, alt=alt)

    # ── Blatt 2: Jahrgänge ────────────────────────────────────────────────────
    ws_jg = wb.create_sheet("Jahrgänge")
    headers = ["Jahrgang", "Anzahl Klassen", "Soll-Stunden/Woche", "Klassen-Buchstaben"]
    for col, h in enumerate(headers, 1):
        style_header(ws_jg.cell(row=1, column=col, value=h))
    for w, width in zip(range(1, 5), [12, 16, 18, 24]):
        set_col_width(ws_jg, w, width)

    for r, gd in enumerate(config.grades.grades, 2):
        labels = gd.class_labels or list("abcdefghij"[:gd.num_classes])
        row_vals = [gd.grade, gd.num_classes, gd.weekly_hours_target, ", ".join(labels)]
        alt = (r % 2 == 0)
        for col, val in enumerate(row_vals, 1):
            style_data(ws_jg.cell(row=r, column=col, value=val), alt=alt)

    # ── Blatt 3: Fächer ───────────────────────────────────────────────────────
    ws_fa = wb.create_sheet("Fächer")
    fa_headers = [
        "Fachname", "Kürzel", "Kategorie", "Hauptfach (ja/nein)",
        "Fachraum-Typ", "Doppelstunde Pflicht", "Doppelstunde Bevorzugt",
    ]
    for col, h in enumerate(fa_headers, 1):
        style_header(ws_fa.cell(row=1, column=col, value=h))
    widths_fa = [18, 10, 14, 20, 16, 22, 24]
    for col, w in enumerate(widths_fa, 1):
        set_col_width(ws_fa, col, w)

    # Hinweis in Zeile 2
    note_cell = ws_fa.cell(
        row=2, column=1,
        value="Beispiel-Vorlage — bitte anpassen. "
              "Zeilen können hinzugefügt, geändert oder gelöscht werden.",
    )
    note_cell.font = Font(italic=True, color="888888", size=10)
    ws_fa.merge_cells(start_row=2, start_column=1,
                      end_row=2, end_column=len(fa_headers))

    dv_ja_nein = DataValidation(
        type="list", formula1='"ja,nein"', allow_blank=False, showDropDown=False
    )
    dv_ja_nein.sqref = "D3:D200"
    ws_fa.add_data_validation(dv_ja_nein)
    dv_bool_f = DataValidation(
        type="list", formula1='"ja,nein"', allow_blank=True, showDropDown=False
    )
    dv_bool_f.sqref = "F3:G200"
    ws_fa.add_data_validation(dv_bool_f)

    for r, (name, meta) in enumerate(SUBJECT_METADATA.items(), 3):
        alt = (r % 2 == 0)
        row_vals = [
            name,
            meta["short"],
            meta["category"],
            "ja" if meta["is_hauptfach"] else "nein",
            meta.get("room") or "",
            "ja" if meta.get("double_required") else "nein",
            "ja" if meta.get("double_preferred") else "nein",
        ]
        for col, val in enumerate(row_vals, 1):
            style_data(ws_fa.cell(row=r, column=col, value=val), alt=alt)

    # ── Blatt 4: Stundentafel ─────────────────────────────────────────────────
    ws_st = wb.create_sheet("Stundentafel")
    all_subjects = list(SUBJECT_METADATA.keys())
    grade_nums = sorted(STUNDENTAFEL_GYMNASIUM_SEK1.keys())

    # Hinweis in Zeile 1 (merged)
    note_st = ws_st.cell(
        row=1, column=1,
        value="Wochenstunden pro Klasse und Fach — bitte vollständig ausfüllen. "
              "Die Fächer müssen mit dem Blatt 'Fächer' übereinstimmen.",
    )
    note_st.font = Font(italic=True, color="555555", size=10)
    ws_st.merge_cells(
        start_row=1, start_column=1,
        end_row=1, end_column=len(grade_nums) + 1,
    )

    # Kopfzeile: Jahrgang-Spalten (Zeile 2)
    style_header(ws_st.cell(row=2, column=1, value="Fach"))
    for col, grade in enumerate(grade_nums, 2):
        style_header(ws_st.cell(row=2, column=col, value=f"Jg. {grade}"))

    set_col_width(ws_st, 1, 16)
    for col in range(2, len(grade_nums) + 2):
        set_col_width(ws_st, col, 10)

    for r, subj in enumerate(all_subjects, 3):
        alt = (r % 2 == 0)
        style_data(ws_st.cell(row=r, column=1, value=subj), alt=alt)
        for col, grade in enumerate(grade_nums, 2):
            hours = STUNDENTAFEL_GYMNASIUM_SEK1.get(grade, {}).get(subj, 0)
            cell = ws_st.cell(row=r, column=col, value=hours if hours else "")
            style_data(cell, alt=alt)

    # ── Blatt 5: Lehrkräfte ───────────────────────────────────────────────────
    ws_lk = wb.create_sheet("Lehrkräfte")
    lk_headers = [
        "Name (Nachname, Vorname)", "Kürzel", "Fächer (kommagetrennt)",
        "Deputat", "Teilzeit", "Sperrzeiten (z.B. Mo1,Di3,Fr5)",
        "Wunschtage (z.B. Mo,Fr)", "Max Std/Tag", "Max Springstd/Tag",
        "Sperrslots (Tag:Slot,...)", "Wunsch-frei (Tage)", "Max Springstd/Woche",
    ]
    for col, h in enumerate(lk_headers, 1):
        style_header(ws_lk.cell(row=1, column=col, value=h))

    widths_lk = [28, 10, 32, 10, 10, 26, 22, 12, 16, 22, 18, 18]
    for col, w in enumerate(widths_lk, 1):
        set_col_width(ws_lk, col, w)

    # Beispielzeile (kursiv)
    example_row = [
        "Müller, Hans", "MÜL", "Mathematik, Physik",
        26, "nein", "Mi5", "Fr", 6, 2,
        "Mo:3,Fr:6", "Fr", 5,
    ]
    for col, val in enumerate(example_row, 1):
        style_example(ws_lk.cell(row=2, column=col, value=val))
    ws_lk.cell(row=2, column=1).comment = None

    dv_teilzeit = DataValidation(
        type="list",
        formula1='"ja,nein"',
        allow_blank=False,
        showDropDown=False,
    )
    dv_teilzeit.sqref = "E3:E200"
    ws_lk.add_data_validation(dv_teilzeit)

    dv_max_h = DataValidation(
        type="whole", operator="between",
        formula1="1", formula2="8",
        allow_blank=True,
    )
    dv_max_h.sqref = "H3:H200"
    ws_lk.add_data_validation(dv_max_h)

    dv_max_g = DataValidation(
        type="whole", operator="between",
        formula1="0", formula2="4",
        allow_blank=True,
    )
    dv_max_g.sqref = "I3:I200"
    ws_lk.add_data_validation(dv_max_g)

    dv_max_gw = DataValidation(
        type="whole", operator="between",
        formula1="0", formula2="20",
        allow_blank=True,
    )
    dv_max_gw.sqref = "L3:L200"
    ws_lk.add_data_validation(dv_max_gw)

    # ── Blatt 6: Fachräume ────────────────────────────────────────────────────
    ws_fr = wb.create_sheet("Fachräume")
    fr_headers = ["Raumtyp (intern)", "Anzeigename", "Anzahl"]
    for col, h in enumerate(fr_headers, 1):
        style_header(ws_fr.cell(row=1, column=col, value=h))
    set_col_width(ws_fr, 1, 18)
    set_col_width(ws_fr, 2, 22)
    set_col_width(ws_fr, 3, 10)

    for r, room_def in enumerate(config.rooms.special_rooms, 2):
        alt = (r % 2 == 0)
        row_vals = [room_def.room_type, room_def.display_name, room_def.count]
        for col, val in enumerate(row_vals, 1):
            style_data(ws_fr.cell(row=r, column=col, value=val), alt=alt)

    # Beispielzeile
    example_fr = ["sport", "Sporthalle", 2]
    for col, val in enumerate(example_fr, 1):
        style_example(ws_fr.cell(row=len(config.rooms.special_rooms) + 2, column=col, value=val))

    # ── Blatt 7: Kopplungen ───────────────────────────────────────────────────
    ws_kp = wb.create_sheet("Kopplungen")
    kp_headers = [
        "ID", "Typ (reli_ethik/wpf)", "Beteiligte Klassen (kommagetrennt)",
        "Gruppen (Name:Fach:Std, kommagetrennt)", "Stunden/Woche", "Klassenübergreifend",
    ]
    for col, h in enumerate(kp_headers, 1):
        style_header(ws_kp.cell(row=1, column=col, value=h))
    widths_kp = [14, 20, 34, 42, 14, 18]
    for col, w in enumerate(widths_kp, 1):
        set_col_width(ws_kp, col, w)

    # Beispiel-Kopplung
    ex_kp = [
        "reli_5", "reli_ethik", "5a,5b,5c,5d,5e,5f",
        "evangelisch:Religion:2,katholisch:Religion:2,ethik:Ethik:2",
        2, "ja",
    ]
    for col, val in enumerate(ex_kp, 1):
        style_example(ws_kp.cell(row=2, column=col, value=val))

    # ── Speichern ─────────────────────────────────────────────────────────────
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


# ─── IMPORT ───────────────────────────────────────────────────────────────────

class ExcelImporter:
    """Importiert Schuldaten aus einer Excel-Vorlage."""

    def __init__(self, path: Path, config: SchoolConfig) -> None:
        self.path = Path(path)
        self.config = config
        self._wb = None
        self._known_subjects = list(SUBJECT_METADATA.keys())
        self._errors: list[str] = []
        self._warnings: list[str] = []

    def _open(self):
        try:
            import openpyxl
            self._wb = openpyxl.load_workbook(
                str(self.path), read_only=True, data_only=True
            )
        except FileNotFoundError:
            raise ExcelImportError(f"Datei nicht gefunden: {self.path}")
        except Exception as e:
            raise ExcelImportError(f"Fehler beim Öffnen der Excel-Datei: {e}")

    def _get_sheet(self, name: str):
        if self._wb is None:
            self._open()
        for sn in self._wb.sheetnames:
            if sn.strip().lower() == name.strip().lower():
                return self._wb[sn]
        return None

    def _sheet_rows(self, sheet) -> list[dict]:
        """Tabellenblatt → Liste von Dicts (erste Zeile = Header)."""
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [
            str(h).strip().lower() if h is not None else f"col_{i}"
            for i, h in enumerate(rows[0])
        ]
        result = []
        for row in rows[1:]:
            if all(v is None or v == "" for v in row):
                continue
            result.append({
                headers[i]: (str(v).strip() if v is not None else "")
                for i, v in enumerate(row)
                if i < len(headers)
            })
        return result

    def _parse_subject(self, raw: str, row_id: str) -> Optional[str]:
        """Normalisiert und validiert einen Fachnamen. Fuzzy-matching bei Tippfehlern."""
        name = raw.strip()
        if not name:
            return None
        if name in self._known_subjects:
            return name
        match = _fuzzy_subject(name, self._known_subjects)
        if match:
            self._warnings.append(
                f"{row_id}: Fach '{name}' unbekannt → meinten Sie '{match}'? "
                f"Wird als '{match}' importiert."
            )
            return match
        self._errors.append(
            f"{row_id}: Unbekanntes Fach '{name}'. "
            f"Ähnliche Fächer: {', '.join(difflib.get_close_matches(name, self._known_subjects, n=3, cutoff=0.4)) or 'keine'}"
        )
        return None

    # ── Fächer ──────────────────────────────────────────────────────────────

    def import_subjects(self) -> list[Subject]:
        """Importiert Fächer aus Blatt 'Fächer'.

        Fallback auf SUBJECT_METADATA wenn das Blatt komplett fehlt.
        Unvollständige Zeilen werden mit Warnung übersprungen.
        """
        sheet = self._get_sheet("Fächer")
        if sheet is None:
            # Blatt fehlt komplett → NRW-Defaults
            return [
                Subject(
                    name=name,
                    short_name=meta["short"],
                    category=meta["category"],
                    is_hauptfach=meta["is_hauptfach"],
                    requires_special_room=meta["room"],
                    double_lesson_required=meta["double_required"],
                    double_lesson_preferred=meta["double_preferred"],
                )
                for name, meta in SUBJECT_METADATA.items()
            ]

        rows = self._sheet_rows(sheet)
        # Hinweis-Zeile (kursiv-Text) herausfiltern
        data_rows = [
            r for r in rows
            if r.get("fachname", "").strip()
            and "vorlage" not in r.get("fachname", "").lower()
            and "beispiel" not in r.get("fachname", "").lower()
        ]

        if not data_rows:
            self._warnings.append(
                "Blatt 'Fächer' ist vorhanden aber leer → "
                "Fallback auf NRW-Standard-Fächer."
            )
            return [
                Subject(
                    name=name,
                    short_name=meta["short"],
                    category=meta["category"],
                    is_hauptfach=meta["is_hauptfach"],
                    requires_special_room=meta["room"],
                    double_lesson_required=meta["double_required"],
                    double_lesson_preferred=meta["double_preferred"],
                )
                for name, meta in SUBJECT_METADATA.items()
            ]

        subjects = []
        for i, row in enumerate(data_rows, 3):
            name = row.get("fachname", "").strip()
            if not name:
                continue
            short = row.get("kürzel", row.get("kurzel", name[:2])).strip() or name[:2]
            category = row.get("kategorie", "sonstige").strip() or "sonstige"
            hf_raw = row.get("hauptfach (ja/nein)", "nein").strip().lower()
            is_hf = hf_raw in ("ja", "yes", "true", "1", "x")
            room_type = row.get("fachraum-typ", "").strip() or None
            dr_raw = row.get("doppelstunde pflicht", "nein").strip().lower()
            dp_raw = row.get("doppelstunde bevorzugt", "nein").strip().lower()
            subjects.append(Subject(
                name=name,
                short_name=short,
                category=category,
                is_hauptfach=is_hf,
                requires_special_room=room_type,
                double_lesson_required=dr_raw in ("ja", "yes", "true", "1", "x"),
                double_lesson_preferred=dp_raw in ("ja", "yes", "true", "1", "x"),
            ))

        if not subjects:
            self._warnings.append(
                "Blatt 'Fächer' enthält keine gültigen Einträge → "
                "Fallback auf NRW-Standard-Fächer."
            )
            return [
                Subject(
                    name=name,
                    short_name=meta["short"],
                    category=meta["category"],
                    is_hauptfach=meta["is_hauptfach"],
                    requires_special_room=meta["room"],
                    double_lesson_required=meta["double_required"],
                    double_lesson_preferred=meta["double_preferred"],
                )
                for name, meta in SUBJECT_METADATA.items()
            ]

        # Aktualisiere bekannte Fächer für _parse_subject
        self._known_subjects = [s.name for s in subjects]
        return subjects

    # ── Stundentafel ────────────────────────────────────────────────────────

    def import_stundentafel(
        self, known_subjects: list[str]
    ) -> dict[int, dict[str, int]]:
        """Importiert die Stundentafel aus Blatt 'Stundentafel'.

        Gibt {Jahrgang: {Fach: Stunden}} zurück.
        Fallback auf STUNDENTAFEL_GYMNASIUM_SEK1 wenn das Blatt fehlt.
        Bei vorhandenem Blatt werden NUR die eingetragenen Werte verwendet —
        fehlende Zellen = 0 Stunden (kein NRW-Inject).
        """
        sheet = self._get_sheet("Stundentafel")
        if sheet is None:
            return dict(STUNDENTAFEL_GYMNASIUM_SEK1)

        # Rohzeilen lesen
        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            return dict(STUNDENTAFEL_GYMNASIUM_SEK1)

        # Hinweis-Zeile überspringen: Blatt hat optionalen Merge in Zeile 1
        # Suche die Kopfzeile (erste Zeile mit "fach" als erstem nicht-leerem Wert)
        header_row_idx = None
        for idx, row in enumerate(all_rows):
            first_val = str(row[0] or "").strip().lower()
            if first_val == "fach":
                header_row_idx = idx
                break

        if header_row_idx is None:
            self._warnings.append(
                "Stundentafel-Blatt hat keine 'Fach'-Kopfzeile → "
                "Fallback auf NRW-Stundentafel."
            )
            return dict(STUNDENTAFEL_GYMNASIUM_SEK1)

        header = all_rows[header_row_idx]
        # Erste Spalte = Fachname, Rest = Jahrgänge (z.B. "Jg. 5" oder "5")
        grade_cols: dict[int, int] = {}  # col_index → grade_number
        for col_idx, cell_val in enumerate(header[1:], 1):
            if cell_val is None:
                continue
            raw = str(cell_val).strip().lower().replace("jg.", "").replace("jg", "").strip()
            try:
                grade = int(float(raw))
                grade_cols[col_idx] = grade
            except ValueError:
                pass

        if not grade_cols:
            self._warnings.append(
                "Stundentafel-Blatt hat keine Jahrgangs-Spalten → "
                "Fallback auf NRW-Stundentafel."
            )
            return dict(STUNDENTAFEL_GYMNASIUM_SEK1)

        result: dict[int, dict[str, int]] = {g: {} for g in grade_cols.values()}

        for row in all_rows[header_row_idx + 1:]:
            if all(v is None for v in row):
                continue
            subj_raw = str(row[0] or "").strip()
            if not subj_raw or subj_raw.lower() == "fach":
                continue
            # Fach validieren (fuzzy)
            subj = subj_raw if subj_raw in known_subjects else None
            if subj is None:
                match = _fuzzy_subject(subj_raw, known_subjects)
                if match:
                    subj = match
                else:
                    self._warnings.append(
                        f"Stundentafel: Unbekanntes Fach '{subj_raw}' — Zeile übersprungen."
                    )
                    continue
            for col_idx, grade in grade_cols.items():
                cell_val = row[col_idx] if col_idx < len(row) else None
                try:
                    hours = int(float(str(cell_val))) if cell_val not in (None, "") else 0
                except (ValueError, TypeError):
                    hours = 0
                if hours > 0:
                    result[grade][subj] = hours

        return result

    # ── Zeitraster ──────────────────────────────────────────────────────────

    def import_time_grid(self):
        """Importiert das Zeitraster aus Blatt 'Zeitraster' und überschreibt Config."""
        from config.schema import LessonSlot, TimeGridConfig

        sheet = self._get_sheet("Zeitraster")
        if sheet is None:
            return  # Optional, kein Fehler

        rows = self._sheet_rows(sheet)
        new_slots = []
        for i, row in enumerate(rows, 2):
            slot_raw = row.get("slot-nr", row.get("slot_nr", ""))
            start = row.get("beginn", row.get("start", ""))
            end = row.get("ende", row.get("end", ""))
            sii_raw = row.get("sii-only", row.get("sii_only", "nein")).lower()

            try:
                slot_num = int(float(slot_raw)) if slot_raw else None
                if slot_num is None:
                    continue
                new_slots.append(LessonSlot(
                    slot_number=slot_num,
                    start_time=start or "00:00",
                    end_time=end or "00:00",
                    is_sek2_only=sii_raw in ("ja", "yes", "true", "1", "x"),
                ))
            except Exception as e:
                self._warnings.append(f"Zeitraster Zeile {i}: Übersprungen ({e})")

        if new_slots:
            # Überschreibt lesson_slots, behält pauses/double_blocks
            self.config = self.config.model_copy(update={
                "time_grid": self.config.time_grid.model_copy(
                    update={"lesson_slots": new_slots}
                )
            })

    # ── Lehrkräfte ─────────────────────────────────────────────────────────

    def import_teachers(self) -> list[Teacher]:
        sheet = self._get_sheet("Lehrkräfte")
        if sheet is None:
            raise ExcelImportError(
                "Tabellenblatt 'Lehrkräfte' nicht gefunden."
            )
        rows = self._sheet_rows(sheet)
        teachers = []
        used_ids: set[str] = set()
        tc = self.config.teachers

        for i, row in enumerate(rows, 2):
            # Beispielzeilen (kursiv-Marker) überspringen anhand von Kürzel = MÜL
            abbr = row.get("kürzel", row.get("kurzel", "")).strip().upper()
            name = row.get("name (nachname, vorname)", row.get("name", "")).strip()

            if not abbr or abbr == "MÜL":
                continue  # Beispielzeile
            if abbr in used_ids:
                self._errors.append(f"Zeile {i}: Doppeltes Kürzel '{abbr}'")
                continue
            used_ids.add(abbr)

            # Fächer — neues Format (kommagetrennt) hat Vorrang vor altem Format (Fach 1/2/3)
            subjects = []
            komma_raw = row.get("fächer (kommagetrennt)", row.get("faecher (kommagetrennt)", "")).strip()
            if komma_raw:
                # Neues Format
                for item in komma_raw.split(","):
                    item = item.strip()
                    if item:
                        s = self._parse_subject(item, f"Zeile {i}, Kürzel {abbr}")
                        if s:
                            subjects.append(s)
            else:
                # Altes Format (Rückwärtskompatibilität)
                for fach_key in ["fach 1", "fach1", "fach 2", "fach2", "fach 3", "fach3"]:
                    raw = row.get(fach_key, "").strip()
                    if raw:
                        s = self._parse_subject(raw, f"Zeile {i}, Kürzel {abbr}")
                        if s:
                            subjects.append(s)

            # Deputat
            dep_raw = row.get("deputat", "").strip()
            try:
                deputat = int(float(dep_raw)) if dep_raw else tc.vollzeit_deputat
            except ValueError:
                self._warnings.append(f"Zeile {i}: Ungültiges Deputat '{dep_raw}' → {tc.vollzeit_deputat}h")
                deputat = tc.vollzeit_deputat

            # Teilzeit
            tz_raw = row.get("teilzeit", "nein").strip().lower()
            is_teilzeit = tz_raw in ("ja", "yes", "true", "1", "x")

            # Sperrzeiten — neues Format (Tag:Slot) hat Vorrang vor altem Format
            sperrslots_raw = row.get(
                "sperrslots (tag:slot,...)", row.get("sperrslots", "")
            ).strip()
            if sperrslots_raw:
                unavailable = _parse_sperrslots(sperrslots_raw)
            else:
                blocked_raw = row.get("sperrzeiten (z.b. mo1,di3,fr5)",
                                      row.get("sperrzeiten", ""))
                unavailable = _parse_blocked_slots(blocked_raw)

            # Wunschtage — neues Format (Leerzeichen/Komma) hat Vorrang
            wunsch_frei_raw = row.get(
                "wunsch-frei (tage)", row.get("wunsch-frei", "")
            ).strip()
            if wunsch_frei_raw:
                free_days = _parse_free_days_flexible(wunsch_frei_raw)
            else:
                wishes_raw = row.get("wunschtage (z.b. mo,fr)",
                                     row.get("wunschtage", ""))
                free_days = _parse_free_days(wishes_raw)

            # Max Std/Tag
            max_h_raw = row.get("max std/tag", row.get("max_std_tag", "")).strip()
            try:
                max_h = int(float(max_h_raw)) if max_h_raw else tc.max_hours_per_day
            except ValueError:
                max_h = tc.max_hours_per_day

            # Max Springstd/Tag
            max_g_raw = row.get("max springstd/tag", row.get("max_springstd", "")).strip()
            try:
                max_g = int(float(max_g_raw)) if max_g_raw else tc.max_gaps_per_day
            except ValueError:
                max_g = tc.max_gaps_per_day

            # Max Springstd/Woche (pro-Lehrer)
            max_gw_raw = row.get("max springstd/woche", "").strip()
            try:
                max_gw = int(float(max_gw_raw)) if max_gw_raw else tc.max_gaps_per_week
            except ValueError:
                max_gw = tc.max_gaps_per_week

            deputat_max = deputat + tc.deputat_max_buffer
            deputat_min = max(1, round(deputat_max * tc.deputat_min_fraction))
            teachers.append(Teacher(
                id=abbr,
                name=name,
                subjects=subjects,
                deputat_max=deputat_max,
                deputat_min=deputat_min,
                is_teilzeit=is_teilzeit,
                unavailable_slots=unavailable,
                preferred_free_days=free_days,
                max_hours_per_day=max_h,
                max_gaps_per_day=max_g,
                max_gaps_per_week=max_gw,
            ))

        if not teachers:
            raise ExcelImportError(
                "Keine Lehrkräfte gefunden. Bitte Tabellenblatt 'Lehrkräfte' prüfen."
            )
        return teachers

    # ── Fachräume ──────────────────────────────────────────────────────────

    def import_rooms(self) -> list[Room]:
        sheet = self._get_sheet("Fachräume")
        if sheet is None:
            return []  # Optional

        rows = self._sheet_rows(sheet)
        rooms = []
        for i, row in enumerate(rows, 2):
            rtype = row.get("raumtyp (intern)", row.get("raumtyp", "")).strip().lower()
            name = row.get("anzeigename", row.get("name", "")).strip()
            count_raw = row.get("anzahl", "1").strip()

            if not rtype or rtype in ("raumtyp", "beispiel"):
                continue  # Header/Beispiel überspringen

            try:
                count = int(float(count_raw)) if count_raw else 1
            except ValueError:
                count = 1

            prefix = rtype[:2].upper()
            for idx in range(1, count + 1):
                rooms.append(Room(
                    id=f"{prefix}{idx}",
                    room_type=rtype,
                    name=f"{name} {idx}" if count > 1 else name,
                ))

        return rooms

    # ── Klassen aus Jahrgänge-Blatt ────────────────────────────────────────

    def import_classes(
        self,
        stundentafel: Optional[dict[int, dict[str, int]]] = None,
    ) -> list[SchoolClass]:
        """Importiert Klassen aus dem 'Jahrgänge'-Blatt.

        Args:
            stundentafel: {Jahrgang: {Fach: Stunden}}.
                          Wenn None, wird STUNDENTAFEL_GYMNASIUM_SEK1 verwendet.
        """
        if stundentafel is None:
            stundentafel = STUNDENTAFEL_GYMNASIUM_SEK1

        sheet = self._get_sheet("Jahrgänge")
        if sheet is None:
            return []

        rows = self._sheet_rows(sheet)
        classes = []
        sek1_max = self.config.time_grid.sek1_max_slot

        for i, row in enumerate(rows, 2):
            grade_raw = row.get("jahrgang", "").strip()
            num_raw = row.get("anzahl klassen", row.get("klassen", "")).strip()

            if not grade_raw:
                continue
            try:
                grade = int(float(grade_raw))
            except ValueError:
                self._warnings.append(
                    f"Jahrgänge Zeile {i}: Ungültiger Jahrgang '{grade_raw}'"
                )
                continue

            try:
                num_classes = int(float(num_raw)) if num_raw else 1
            except ValueError:
                num_classes = 1

            curriculum = {
                f: h
                for f, h in stundentafel.get(grade, {}).items()
                if h > 0
            }

            labels = list("abcdefghij"[:num_classes])
            for label in labels:
                classes.append(SchoolClass(
                    id=f"{grade}{label}",
                    grade=grade,
                    label=label,
                    curriculum=curriculum.copy(),
                    max_slot=sek1_max,
                ))

        return classes

    # ── Kopplungen ─────────────────────────────────────────────────────────

    def import_couplings(self) -> list[Coupling]:
        sheet = self._get_sheet("Kopplungen")
        if sheet is None:
            return []

        rows = self._sheet_rows(sheet)
        couplings = []

        for i, row in enumerate(rows, 2):
            cid = row.get("id", "").strip()
            ctype = row.get("typ (reli_ethik/wpf)", row.get("typ", "")).strip().lower()
            classes_raw = row.get("beteiligte klassen (kommagetrennt)",
                                  row.get("klassen", "")).strip()
            groups_raw = row.get("gruppen (name:fach:std, kommagetrennt)",
                                 row.get("gruppen", "")).strip()
            hours_raw = row.get("stunden/woche", row.get("stunden", "2")).strip()
            cross_raw = row.get("klassenübergreifend", "ja").strip().lower()

            if not cid or cid.startswith("id"):
                continue  # Header/Beispiel

            class_ids = [c.strip() for c in classes_raw.split(",") if c.strip()]

            # Gruppen parsen: "Name:Fach:Std,..."
            groups = []
            for gpart in groups_raw.split(","):
                parts = gpart.strip().split(":")
                if len(parts) >= 2:
                    gname = parts[0].strip()
                    gsubj = parts[1].strip() if len(parts) > 1 else ""
                    ghours_raw = parts[2].strip() if len(parts) > 2 else ""
                    try:
                        ghours = int(ghours_raw) if ghours_raw else 2
                    except ValueError:
                        ghours = 2
                    groups.append(CouplingGroup(
                        group_name=gname,
                        subject=gsubj,
                        hours_per_week=ghours,
                    ))

            try:
                hours = int(float(hours_raw)) if hours_raw else 2
            except ValueError:
                hours = 2

            cross = cross_raw not in ("nein", "no", "false", "0")

            if not groups:
                self._warnings.append(f"Kopplung Zeile {i}: Keine Gruppen definiert für '{cid}'")

            couplings.append(Coupling(
                id=cid,
                coupling_type=ctype or "wpf",
                involved_class_ids=class_ids,
                groups=groups,
                hours_per_week=hours,
                cross_class=cross,
            ))

        return couplings

    # ── Vollständiger Import ───────────────────────────────────────────────

    def import_all(self) -> tuple[SchoolData, FeasibilityReport]:
        """Importiert alle Daten → SchoolData + FeasibilityReport."""
        self._open()
        self._errors = []
        self._warnings = []

        # Zeitraster optional überschreiben
        self.import_time_grid()

        # Fächer aus 'Fächer'-Blatt (Fallback: SUBJECT_METADATA)
        subjects = self.import_subjects()
        known_names = [s.name for s in subjects]

        # Stundentafel aus 'Stundentafel'-Blatt (Fallback: NRW-Defaults)
        stundentafel = self.import_stundentafel(known_names)

        # Räume
        try:
            rooms = self.import_rooms()
        except ExcelImportError as e:
            self._errors.append(f"Fachräume: {e}")
            rooms = []

        # Klassen
        try:
            classes = self.import_classes(stundentafel=stundentafel)
        except ExcelImportError as e:
            self._errors.append(f"Klassen: {e}")
            classes = []

        # Lehrkräfte
        try:
            teachers = self.import_teachers()
        except ExcelImportError as e:
            self._errors.append(f"Lehrkräfte: {e}")
            teachers = []

        # Kopplungen
        try:
            couplings = self.import_couplings()
        except ExcelImportError as e:
            self._warnings.append(f"Kopplungen: {e}")
            couplings = []

        if self._errors:
            # Kritische Fehler: Report zurückgeben, kein SchoolData
            report = FeasibilityReport(
                is_feasible=False,
                errors=self._errors,
                warnings=self._warnings,
            )
            raise ExcelImportError(
                f"Import mit {len(self._errors)} Fehlern:\n"
                + "\n".join(f"  • {e}" for e in self._errors)
            )

        school_data = SchoolData(
            subjects=subjects,
            rooms=rooms,
            classes=classes,
            teachers=teachers,
            couplings=couplings,
            config=self.config,
        )

        # Machbarkeits-Check nach Import
        feasibility = school_data.validate_feasibility()
        feasibility = FeasibilityReport(
            is_feasible=feasibility.is_feasible,
            errors=feasibility.errors,
            warnings=feasibility.warnings + self._warnings,
        )

        return school_data, feasibility


def import_from_excel(
    path: Path, config: SchoolConfig
) -> tuple[SchoolData, FeasibilityReport]:
    """Importiert Schuldaten aus einer Excel-Vorlage.

    Args:
        path:   Pfad zur Excel-Datei (.xlsx)
        config: Basis-Konfiguration (kann durch Zeitraster-Blatt überschrieben werden)

    Returns:
        (SchoolData, FeasibilityReport)

    Raises:
        ExcelImportError: Bei kritischen Import-Fehlern.
    """
    importer = ExcelImporter(path, config)
    return importer.import_all()


# ─── CSV-IMPORTER ──────────────────────────────────────────────────────────────

# Dateiname → Sheet-Name Mapping (case-insensitive, Umlaute normalisiert)
_CSV_SHEET_MAP: dict[str, str] = {
    "lehrkraefte": "Lehrkräfte",
    "lehrkräfte": "Lehrkräfte",
    "lehrkraefte": "Lehrkräfte",
    "stundentafel": "Stundentafel",
    "jahrgaenge": "Jahrgänge",
    "jahrgänge": "Jahrgänge",
    "faecher": "Fächer",
    "fächer": "Fächer",
    "fachraeume": "Fachräume",
    "fachräume": "Fachräume",
    "kopplungen": "Kopplungen",
    "zeitraster": "Zeitraster",
}


def _normalize_filename(name: str) -> str:
    """Normalisiert Dateinamen für das Mapping (Kleinbuchstaben, kein Suffix)."""
    import unicodedata
    name = name.lower()
    # Umlaute belassen — Mapping hat beide Varianten
    return name


class CsvImporter(ExcelImporter):
    """Importiert Schuldaten aus CSV-Dateien.

    Akzeptiert:
    - Ein Verzeichnis mit CSV-Dateien benannt nach ihren Blättern
      (z.B. Lehrkraefte.csv, Stundentafel.csv, ...)
    - Eine einzelne .csv-Datei → wird als 'Lehrkräfte'-Blatt behandelt
    """

    def __init__(self, path: Path, config: SchoolConfig) -> None:
        super().__init__(path, config)
        self._csv_sheets: dict[str, list[dict]] = {}

    def _open(self) -> None:
        import csv

        path = Path(self.path)
        if path.is_dir():
            for csv_file in sorted(path.glob("*.csv")):
                stem = _normalize_filename(csv_file.stem)
                sheet_name = _CSV_SHEET_MAP.get(stem)
                if sheet_name is None:
                    # Try partial match
                    for key, val in _CSV_SHEET_MAP.items():
                        if key in stem or stem in key:
                            sheet_name = val
                            break
                if sheet_name is None:
                    sheet_name = csv_file.stem  # Fallback: Dateiname als Sheet-Name
                with open(csv_file, encoding="utf-8-sig", newline="") as f:
                    reader = csv.DictReader(f)
                    rows = [
                        {k.strip(): (v.strip() if v else "") for k, v in row.items()}
                        for row in reader
                    ]
                self._csv_sheets[sheet_name] = rows
        elif path.suffix.lower() == ".csv":
            with open(path, encoding="utf-8-sig", newline="") as f:
                import csv as _csv
                reader = _csv.DictReader(f)
                rows = [
                    {k.strip(): (v.strip() if v else "") for k, v in row.items()}
                    for row in reader
                ]
            self._csv_sheets["Lehrkräfte"] = rows
        else:
            raise ExcelImportError(
                f"Unbekanntes Dateiformat: {path}. "
                "Erwartet: .xlsx, .csv oder Verzeichnis mit CSV-Dateien."
            )

    def _get_sheet(self, name: str):
        if not self._csv_sheets:
            self._open()
        # Suche case-insensitive
        for sn, rows in self._csv_sheets.items():
            if sn.strip().lower() == name.strip().lower():
                return _CsvSheetProxy(rows)
        return None


class _CsvSheetProxy:
    """Adapter, der eine Liste von Dicts als Sheet-ähnliches Objekt bereitstellt.

    Kompatibel mit ExcelImporter._sheet_rows() indem iter_rows() emuliert wird.
    """

    def __init__(self, rows: list[dict]) -> None:
        self._rows = rows

    def iter_rows(self, values_only: bool = True):
        if not self._rows:
            return iter([])
        headers = list(self._rows[0].keys())
        yield tuple(headers)
        for row in self._rows:
            yield tuple(row.get(h, "") for h in headers)


def import_from_csv(
    path: Path, config: SchoolConfig
) -> tuple[SchoolData, FeasibilityReport]:
    """Importiert Schuldaten aus CSV-Datei(en).

    Args:
        path:   Pfad zu einer .csv-Datei oder einem Verzeichnis mit CSV-Dateien
        config: Basis-Konfiguration

    Returns:
        (SchoolData, FeasibilityReport)

    Raises:
        ExcelImportError: Bei kritischen Import-Fehlern.
    """
    importer = CsvImporter(path, config)
    return importer.import_all()
